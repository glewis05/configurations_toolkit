"""
Access Excel Formatter - Export user access data to Excel

PURPOSE: Generate formatted Excel exports of user access data
         and compliance reports for auditor review

R EQUIVALENT: Like openxlsx or writexl packages for creating
formatted Excel workbooks with compliance-focused layouts

AVIATION ANALOGY: Like generating crew qualification reports -
formatted spreadsheets showing who is rated for what aircraft,
their medical status, and recurrent training dates, all in a
format that FAA inspectors can easily review

AUTHOR: Glen Lewis
DATE: 2025

REPORT FORMATS:
    Access Report:
        Tab 1: Summary - Counts and overview statistics
        Tab 2: Active Access - All current access grants
        Tab 3: Review Status - Overdue and upcoming reviews
        Tab 4: Training Status - Current and expired training
        Tab 5: Audit Trail - Recent access changes

    Review Worksheet:
        Single sheet with dropdown for reviewer decisions
        Can be exported, filled in, and imported back
"""

import os
from pathlib import Path
from datetime import datetime, date
from typing import Dict, List, Optional, Any

# openpyxl for Excel generation
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not installed. Run: pip install openpyxl")

import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from managers.access_manager import AccessManager


class AccessExcelFormatter:
    """
    PURPOSE: Export user access data and compliance reports to Excel

    This class creates formatted Excel workbooks for:
    - Access reports (who has access to what)
    - Review worksheets (for conducting periodic reviews)
    - Compliance reports (for auditor review)

    R EQUIVALENT:
        In R, you'd use openxlsx to create workbooks with multiple sheets,
        styled headers, and data validation. This class wraps those operations
        into methods like export_access_report() and export_review_worksheet().

    PARAMETERS:
        access_manager: AccessManager instance for data access

    EXAMPLE:
        am = AccessManager()
        formatter = AccessExcelFormatter(am)
        formatter.export_access_report(access_data, "access_report.xlsx")

    FORMATTING FEATURES:
    - Professional color scheme matching ConfigExcelFormatter
    - Summary sheets with key metrics
    - Freeze panes for easy navigation
    - Conditional formatting for status indicators
    - Data validation dropdowns for review worksheets
    """

    # =========================================================================
    # STYLE DEFINITIONS - Match ConfigExcelFormatter for consistency
    # =========================================================================

    # Primary header style (dark blue with white text)
    HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Alternating row colors for readability
    ALT_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # Status colors
    PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    PASS_FONT = Font(color="006100", bold=True)

    FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    FAIL_FONT = Font(color="9C0006", bold=True)

    WARNING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    WARNING_FONT = Font(color="9C5700", bold=True)

    # Section/category headers
    SECTION_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    SECTION_FONT = Font(bold=True, size=11)

    # Title styling
    TITLE_FONT = Font(size=18, bold=True, color="2F5496")
    SUBTITLE_FONT = Font(size=12, bold=True, color="595959")

    # Borders
    THIN_BORDER = Border(
        left=Side(style='thin', color='B4B4B4'),
        right=Side(style='thin', color='B4B4B4'),
        top=Side(style='thin', color='B4B4B4'),
        bottom=Side(style='thin', color='B4B4B4')
    )

    # Default cell alignment
    DEFAULT_ALIGNMENT = Alignment(vertical='center', wrap_text=True)
    CENTER_ALIGNMENT = Alignment(horizontal='center', vertical='center')

    def __init__(self, access_manager: AccessManager):
        """
        Initialize with an AccessManager instance.

        PURPOSE: Set up the formatter with database access

        PARAMETERS:
            access_manager: AccessManager instance for data queries

        RAISES:
            ImportError: If openpyxl is not installed
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl required. Install with: pip install openpyxl")

        self.am = access_manager

    def export_access_report(
        self,
        access_data: Dict[str, Any],
        output_path: str
    ) -> str:
        """
        Create Excel report with multiple tabs for access data.

        PURPOSE: Export access list with summary, details, and audit trail

        PARAMETERS:
            access_data: Output from ComplianceReports.access_list_report()
            output_path: Where to save the Excel file

        RETURNS:
            str: Path to generated file

        TABS:
            1. Summary - Key metrics and status overview
            2. Active Access - All current access grants
            3. Review Status - Access needing review
            4. Training - Training status by user
            5. Audit Trail - Recent access changes
        """
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Create sheets
        self._create_summary_sheet(wb, access_data)
        self._create_access_list_sheet(wb, access_data)

        # Ensure output directory exists
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb.save(str(output_path))
        return str(output_path)

    def export_review_worksheet(
        self,
        access_list: List[Dict],
        output_path: str,
        scope_description: str = None
    ) -> str:
        """
        Create worksheet for conducting access reviews.

        PURPOSE: Generate a form that reviewers can fill in and import back

        R EQUIVALENT:
            Like creating a data entry form in Excel with dropdown validation
            that you can later read back with readxl

        PARAMETERS:
            access_list: List of access grants to review
            output_path: Where to save the Excel file
            scope_description: Description of what's being reviewed

        RETURNS:
            str: Path to generated file

        COLUMNS:
            - User Name (read-only)
            - Email (read-only)
            - Role (read-only)
            - Scope - Program/Clinic/Location (read-only)
            - Granted Date (read-only)
            - Last Review Date (read-only)
            - Decision (dropdown: Certified, Revoked, Modified)
            - Reviewer Notes (free text)

        AVIATION ANALOGY:
            Like a proficiency check form - the examiner marks each item
            as Satisfactory or Unsatisfactory, signs it, and files it.
            Same concept for access reviews.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Access Review"

        # Title row
        ws.merge_cells('A1:H1')
        ws['A1'] = f"Access Review Worksheet - {datetime.now().strftime('%Y-%m-%d')}"
        ws['A1'].font = self.TITLE_FONT

        if scope_description:
            ws.merge_cells('A2:H2')
            ws['A2'] = scope_description
            ws['A2'].font = self.SUBTITLE_FONT

        # Headers start at row 4
        headers = [
            'Access ID', 'User Name', 'Email', 'Role', 'Scope',
            'Granted Date', 'Last Review', 'Decision', 'Reviewer Notes'
        ]

        header_row = 4
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER

        # Data rows
        for i, access in enumerate(access_list, 1):
            row = header_row + i

            # Build scope string
            scope_parts = [access.get('program_name', '')]
            if access.get('clinic_name'):
                scope_parts.append(access['clinic_name'])
            if access.get('location_name'):
                scope_parts.append(access['location_name'])
            scope = ' > '.join(scope_parts)

            values = [
                access.get('access_id', ''),
                access.get('user_name', ''),
                access.get('email', ''),
                access.get('role', ''),
                scope,
                access.get('granted_date', ''),
                access.get('last_review_date', ''),
                '',  # Decision - to be filled in
                ''   # Notes - to be filled in
            ]

            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.THIN_BORDER

                # Alternate row colors
                if i % 2 == 0:
                    cell.fill = self.ALT_ROW_FILL

        # Add dropdown validation for Decision column
        decision_validation = DataValidation(
            type="list",
            formula1='"Certified,Revoked,Modified"',
            allow_blank=True
        )
        decision_validation.error = 'Please select from the dropdown'
        decision_validation.errorTitle = 'Invalid Decision'

        # Apply to all Decision cells
        decision_col = 8  # Column H
        if len(access_list) > 0:
            decision_range = f"H{header_row + 1}:H{header_row + len(access_list)}"
            ws.add_data_validation(decision_validation)
            decision_validation.add(decision_range)

        # Adjust column widths
        column_widths = [10, 25, 30, 15, 40, 12, 12, 12, 40]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # Freeze header row
        ws.freeze_panes = 'A5'

        # Add instructions at bottom
        instruction_row = header_row + len(access_list) + 3
        ws.merge_cells(f'A{instruction_row}:I{instruction_row}')
        ws[f'A{instruction_row}'] = (
            "Instructions: Select a Decision for each row. "
            "'Certified' confirms access is still needed. "
            "'Revoked' will remove access. "
            "'Modified' requires additional action."
        )
        ws[f'A{instruction_row}'].font = Font(italic=True, color="666666")

        # Save
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))

        return str(output_path)

    def export_compliance_report(
        self,
        report_data: Dict[str, Any],
        output_path: str
    ) -> str:
        """
        Export any compliance report to Excel.

        PURPOSE: Create formatted Excel from ComplianceReports output

        PARAMETERS:
            report_data: Output from any ComplianceReports method
            output_path: Where to save the Excel file

        RETURNS:
            str: Path to generated file

        WHY THIS APPROACH:
            Each report type has different data structures, so we dispatch
            to the appropriate formatter based on report_type field.
        """
        report_type = report_data.get('report_type', 'unknown')

        formatters = {
            'access_list': self._format_access_list_report,
            'access_changes': self._format_access_changes_report,
            'review_status': self._format_review_status_report,
            'overdue_reviews': self._format_overdue_reviews_report,
            'training_compliance': self._format_training_report,
            'terminated_audit': self._format_terminated_audit_report,
            'business_associates': self._format_ba_report,
            'segregation_of_duties': self._format_sod_report
        }

        if report_type not in formatters:
            raise ValueError(f"Unknown report type: {report_type}")

        wb = Workbook()
        wb.remove(wb.active)

        formatters[report_type](wb, report_data)

        # Save
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))

        return str(output_path)

    # =========================================================================
    # PRIVATE FORMATTING METHODS
    # =========================================================================

    def _create_summary_sheet(self, wb: Workbook, data: Dict) -> None:
        """Create summary sheet with key metrics."""
        ws = wb.create_sheet("Summary")

        # Title
        ws['A1'] = "Access Report Summary"
        ws['A1'].font = self.TITLE_FONT

        # Report metadata
        ws['A3'] = "Report Date:"
        ws['B3'] = data.get('report_date', datetime.now().isoformat())
        ws['A4'] = "As Of Date:"
        ws['B4'] = data.get('as_of_date', 'Current')

        # Filters applied
        ws['A6'] = "Filters Applied"
        ws['A6'].font = self.SECTION_FONT
        ws['A6'].fill = self.SECTION_FILL

        filters = data.get('filters', {})
        row = 7
        for key, value in filters.items():
            if value:
                ws[f'A{row}'] = f"  {key}:"
                ws[f'B{row}'] = str(value)
                row += 1

        # Summary statistics
        row += 2
        ws[f'A{row}'] = "Summary Statistics"
        ws[f'A{row}'].font = self.SECTION_FONT
        ws[f'A{row}'].fill = self.SECTION_FILL

        summary = data.get('summary', {})
        row += 1
        for key, value in summary.items():
            ws[f'A{row}'] = f"  {key.replace('_', ' ').title()}:"
            ws[f'B{row}'] = value
            row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40

    def _create_access_list_sheet(self, wb: Workbook, data: Dict) -> None:
        """Create detailed access list sheet."""
        ws = wb.create_sheet("Active Access")

        # Headers
        headers = [
            'User ID', 'User Name', 'Email', 'Organization',
            'Role', 'Program', 'Clinic', 'Location',
            'Granted Date', 'Granted By', 'Next Review Due', 'Training Status'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER

        # Data rows
        access_list = data.get('access_list', [])
        for i, access in enumerate(access_list, 1):
            row = i + 1

            # Determine training status summary
            training = access.get('training_summary', {})
            if training.get('expired_count', 0) > 0:
                training_status = 'Expired'
            elif training.get('missing_count', 0) > 0:
                training_status = 'Missing'
            else:
                training_status = 'Current'

            values = [
                access.get('user_id', ''),
                access.get('user_name', ''),
                access.get('email', ''),
                access.get('organization', ''),
                access.get('role', ''),
                access.get('program_name', ''),
                access.get('clinic_name', ''),
                access.get('location_name', ''),
                access.get('granted_date', ''),
                access.get('granted_by', ''),
                access.get('next_review_due', ''),
                training_status
            ]

            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.THIN_BORDER

                # Alternate row colors
                if i % 2 == 0:
                    cell.fill = self.ALT_ROW_FILL

            # Highlight training status
            training_cell = ws.cell(row=row, column=12)
            if training_status == 'Expired':
                training_cell.fill = self.FAIL_FILL
                training_cell.font = self.FAIL_FONT
            elif training_status == 'Missing':
                training_cell.fill = self.WARNING_FILL
                training_cell.font = self.WARNING_FONT
            else:
                training_cell.fill = self.PASS_FILL
                training_cell.font = self.PASS_FONT

        # Adjust column widths
        widths = [15, 25, 30, 20, 15, 20, 20, 25, 12, 20, 12, 12]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # Freeze header row
        ws.freeze_panes = 'A2'

    def _format_access_list_report(self, wb: Workbook, data: Dict) -> None:
        """Format access_list report type."""
        self._create_summary_sheet(wb, data)
        self._create_access_list_sheet(wb, data)

    def _format_access_changes_report(self, wb: Workbook, data: Dict) -> None:
        """Format access_changes report type."""
        # Summary sheet
        ws = wb.create_sheet("Summary")
        ws['A1'] = "Access Changes Report"
        ws['A1'].font = self.TITLE_FONT

        period = data.get('period', {})
        ws['A3'] = f"Period: {period.get('start_date', '')} to {period.get('end_date', '')}"
        ws['A3'].font = self.SUBTITLE_FONT

        summary = data.get('summary', {})
        ws['A5'] = f"New Grants: {summary.get('grants', 0)}"
        ws['A6'] = f"Revocations: {summary.get('revocations', 0)}"
        ws['A7'] = f"Modifications: {summary.get('modifications', 0)}"

        # Grants sheet
        grants = data.get('grants', [])
        if grants:
            self._create_data_sheet(wb, "Grants", grants,
                                    ['user_name', 'email', 'role', 'program_name',
                                     'clinic_name', 'granted_date', 'granted_by', 'grant_reason'])

        # Revocations sheet
        revocations = data.get('revocations', [])
        if revocations:
            self._create_data_sheet(wb, "Revocations", revocations,
                                    ['user_name', 'email', 'role', 'program_name',
                                     'clinic_name', 'revoked_date', 'revoked_by', 'revoke_reason'])

    def _format_review_status_report(self, wb: Workbook, data: Dict) -> None:
        """Format review_status report type."""
        # Summary
        ws = wb.create_sheet("Review Status")
        ws['A1'] = "Access Review Status"
        ws['A1'].font = self.TITLE_FONT

        summary = data.get('summary', {})
        ws['A3'] = f"Total Access Grants: {summary.get('total_access', 0)}"
        ws['A4'] = f"Reviews Current: {summary.get('current', 0)}"
        ws['A5'] = f"Due in 30 Days: {summary.get('due_soon', 0)}"
        ws['A6'] = f"Overdue: {summary.get('overdue', 0)}"
        ws['A7'] = f"Compliance: {summary.get('compliance_percentage', 0)}%"

        # Style the overdue count if any
        if summary.get('overdue', 0) > 0:
            ws['A6'].fill = self.FAIL_FILL
            ws['A6'].font = self.FAIL_FONT

        # Overdue details
        overdue = data.get('overdue', [])
        if overdue:
            self._create_data_sheet(wb, "Overdue Reviews", overdue,
                                    ['user_name', 'email', 'role', 'program_name',
                                     'clinic_name', 'next_review_due', 'last_review_date'])

        # Due soon details
        due_soon = data.get('due_soon', [])
        if due_soon:
            self._create_data_sheet(wb, "Due Soon", due_soon,
                                    ['user_name', 'email', 'role', 'program_name',
                                     'clinic_name', 'next_review_due', 'last_review_date'])

    def _format_overdue_reviews_report(self, wb: Workbook, data: Dict) -> None:
        """Format overdue_reviews report type."""
        ws = wb.create_sheet("Overdue Reviews")
        ws['A1'] = "Overdue Access Reviews"
        ws['A1'].font = self.TITLE_FONT

        summary = data.get('summary', {})
        ws['A3'] = f"Total Overdue: {summary.get('total_overdue', 0)}"
        ws['A4'] = f"Oldest Overdue: {summary.get('oldest_overdue', 'N/A')}"
        ws['A5'] = f"Max Days Overdue: {summary.get('max_days_overdue', 0)}"

        if summary.get('total_overdue', 0) > 0:
            ws['A3'].fill = self.FAIL_FILL
            ws['A3'].font = self.FAIL_FONT

        # Overdue list
        overdue = data.get('overdue_reviews', [])
        if overdue:
            self._create_data_sheet(wb, "Details", overdue,
                                    ['user_name', 'email', 'role', 'program_name',
                                     'clinic_name', 'next_review_due', 'days_overdue'])

    def _format_training_report(self, wb: Workbook, data: Dict) -> None:
        """Format training_compliance report type."""
        # Summary
        ws = wb.create_sheet("Training Summary")
        ws['A1'] = "Training Compliance Report"
        ws['A1'].font = self.TITLE_FONT

        summary = data.get('summary', {})
        ws['A3'] = f"Total Users: {summary.get('total_users', 0)}"
        ws['A4'] = f"Training Current: {summary.get('current', 0)}"
        ws['A5'] = f"Training Expired: {summary.get('expired', 0)}"
        ws['A6'] = f"Training Missing: {summary.get('missing', 0)}"
        ws['A7'] = f"Compliance: {summary.get('compliance_percentage', 0)}%"

        # Expired users
        expired = data.get('expired_users', [])
        if expired:
            self._create_training_sheet(wb, "Expired Training", expired)

        # Missing users
        missing = data.get('missing_users', [])
        if missing:
            self._create_training_sheet(wb, "Missing Training", missing)

    def _format_terminated_audit_report(self, wb: Workbook, data: Dict) -> None:
        """Format terminated_audit report type."""
        ws = wb.create_sheet("Terminated User Audit")
        ws['A1'] = "Terminated User Audit"
        ws['A1'].font = self.TITLE_FONT

        is_compliant = data.get('is_compliant', True)
        summary = data.get('summary', {})

        ws['A3'] = f"Status: {summary.get('status', 'Unknown')}"
        if is_compliant:
            ws['A3'].fill = self.PASS_FILL
            ws['A3'].font = self.PASS_FONT
        else:
            ws['A3'].fill = self.FAIL_FILL
            ws['A3'].font = self.FAIL_FONT

        ws['A5'] = f"Terminated Users with Active Access: {summary.get('terminated_with_access', 0)}"

        # Findings
        findings = data.get('findings', [])
        if findings:
            self._create_data_sheet(wb, "Findings", findings,
                                    ['name', 'email', 'role', 'program_name',
                                     'clinic_name', 'granted_date'])

    def _format_ba_report(self, wb: Workbook, data: Dict) -> None:
        """Format business_associates report type."""
        ws = wb.create_sheet("Business Associates")
        ws['A1'] = "Business Associate Report"
        ws['A1'].font = self.TITLE_FONT

        summary = data.get('summary', {})
        ws['A3'] = f"Total External Users: {summary.get('total_external_users', 0)}"
        ws['A4'] = f"Organizations: {summary.get('organizations', 0)}"
        ws['A5'] = f"Total Access Grants: {summary.get('total_access_grants', 0)}"

        # All external users
        users = data.get('all_external_users', [])
        if users:
            self._create_data_sheet(wb, "External Users", users,
                                    ['name', 'email', 'organization', 'programs', 'access_count'])

    def _format_sod_report(self, wb: Workbook, data: Dict) -> None:
        """Format segregation_of_duties report type."""
        ws = wb.create_sheet("Segregation of Duties")
        ws['A1'] = "Segregation of Duties Report"
        ws['A1'].font = self.TITLE_FONT

        is_compliant = data.get('is_compliant', True)
        summary = data.get('summary', {})

        ws['A3'] = f"Status: {summary.get('status', 'Unknown')}"
        if is_compliant:
            ws['A3'].fill = self.PASS_FILL
            ws['A3'].font = self.PASS_FONT
        else:
            ws['A3'].fill = self.FAIL_FILL
            ws['A3'].font = self.FAIL_FONT

        ws['A5'] = f"Users Checked: {summary.get('users_checked', 0)}"
        ws['A6'] = f"Blocking Violations: {summary.get('blocking_violations', 0)}"
        ws['A7'] = f"Warnings: {summary.get('warnings', 0)}"

        # Violations
        violations = data.get('violations', [])
        if violations:
            ws_v = wb.create_sheet("Violations")
            headers = ['User Name', 'Email', 'Program', 'Conflicting Roles', 'Severity', 'Reason']
            for col, header in enumerate(headers, 1):
                cell = ws_v.cell(row=1, column=col, value=header)
                cell.fill = self.HEADER_FILL
                cell.font = self.HEADER_FONT

            for i, v in enumerate(violations, 1):
                ws_v.cell(row=i+1, column=1, value=v.get('user_name', ''))
                ws_v.cell(row=i+1, column=2, value=v.get('email', ''))
                ws_v.cell(row=i+1, column=3, value=v.get('program_name', ''))
                ws_v.cell(row=i+1, column=4, value=', '.join(v.get('conflicting_roles', [])))
                ws_v.cell(row=i+1, column=5, value=v.get('severity', ''))
                ws_v.cell(row=i+1, column=6, value=v.get('reason', ''))

        # Warnings
        warnings = data.get('warnings', [])
        if warnings:
            ws_w = wb.create_sheet("Warnings")
            headers = ['User Name', 'Email', 'Program', 'Conflicting Roles', 'Severity', 'Reason']
            for col, header in enumerate(headers, 1):
                cell = ws_w.cell(row=1, column=col, value=header)
                cell.fill = self.HEADER_FILL
                cell.font = self.HEADER_FONT

            for i, w in enumerate(warnings, 1):
                ws_w.cell(row=i+1, column=1, value=w.get('user_name', ''))
                ws_w.cell(row=i+1, column=2, value=w.get('email', ''))
                ws_w.cell(row=i+1, column=3, value=w.get('program_name', ''))
                ws_w.cell(row=i+1, column=4, value=', '.join(w.get('conflicting_roles', [])))
                ws_w.cell(row=i+1, column=5, value=w.get('severity', ''))
                ws_w.cell(row=i+1, column=6, value=w.get('reason', ''))

    def _create_data_sheet(
        self,
        wb: Workbook,
        sheet_name: str,
        data: List[Dict],
        columns: List[str]
    ) -> None:
        """
        Create a simple data sheet with headers and rows.

        PURPOSE: Reusable helper for creating tabular data sheets

        PARAMETERS:
            wb: Workbook to add sheet to
            sheet_name: Name for the new sheet
            data: List of dicts to display
            columns: Which dict keys to show as columns
        """
        ws = wb.create_sheet(sheet_name)

        # Format column names for display
        display_headers = [col.replace('_', ' ').title() for col in columns]

        # Headers
        for col, header in enumerate(display_headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER

        # Data rows
        for i, record in enumerate(data, 1):
            for col, key in enumerate(columns, 1):
                value = record.get(key, '')
                cell = ws.cell(row=i+1, column=col, value=value)
                cell.border = self.THIN_BORDER
                if i % 2 == 0:
                    cell.fill = self.ALT_ROW_FILL

        # Auto-width columns (approximate)
        for col in range(1, len(columns) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

        # Freeze header row
        ws.freeze_panes = 'A2'

    def _create_training_sheet(
        self,
        wb: Workbook,
        sheet_name: str,
        users: List[Dict]
    ) -> None:
        """Create a training status sheet for users."""
        ws = wb.create_sheet(sheet_name)

        headers = ['User ID', 'Name', 'Email', 'Training Summary']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.THIN_BORDER

        for i, user in enumerate(users, 1):
            training = user.get('training_summary', {})
            summary_text = []
            if training.get('expired_count', 0) > 0:
                summary_text.append(f"Expired: {training['expired_count']}")
            if training.get('missing_count', 0) > 0:
                summary_text.append(f"Missing: {', '.join(training.get('missing_types', []))}")

            ws.cell(row=i+1, column=1, value=user.get('user_id', ''))
            ws.cell(row=i+1, column=2, value=user.get('name', ''))
            ws.cell(row=i+1, column=3, value=user.get('email', ''))
            ws.cell(row=i+1, column=4, value='; '.join(summary_text))

        # Adjust widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 50

        ws.freeze_panes = 'A2'


# =============================================================================
# MODULE TEST
# =============================================================================

if __name__ == "__main__":
    print("Testing AccessExcelFormatter...")

    if not OPENPYXL_AVAILABLE:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    else:
        print("openpyxl available. Formatter ready to use.")
        print("\nUsage:")
        print("  am = AccessManager()")
        print("  formatter = AccessExcelFormatter(am)")
        print("  formatter.export_review_worksheet(access_list, 'review.xlsx')")
