"""
Configuration Excel Formatter - Export configs to Excel

PURPOSE: Generate formatted Excel exports of configurations
         using a single Configuration Matrix view

R EQUIVALENT: Like openxlsx or writexl packages for creating
formatted Excel workbooks with pivot-table-like views

AVIATION ANALOGY: Like a configuration deviation report -
a single matrix showing what's set at each level (aircraft,
fleet, global) so you can spot differences at a glance

AUTHOR: Glen Lewis
DATE: 2024

MATRIX FORMAT:
    Rows: Configuration keys grouped by category
    Columns: Program Default | Clinic | Location1 | Location2 | ...
    Values: "—" for inherited, actual value for set values
    Highlighting: Yellow (gold) for overrides (differs from parent)
"""

import os
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Any

# openpyxl for Excel generation
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side, NamedStyle
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.formatting.rule import FormulaRule
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not installed. Run: pip install openpyxl")

import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from database.config_manager import ConfigurationManager


class ConfigExcelFormatter:
    """
    PURPOSE: Export configurations to a single Configuration Matrix view

    The matrix shows:
    - Rows grouped by category (e.g., HELPDESK, OPERATIONS, TC SCORING)
    - Columns for each level: Program Default | Clinic | Location1 | Location2...
    - Values: Actual value if set at that level, "—" if inherited
    - Yellow highlighting for overrides (values different from parent level)

    PARAMETERS:
        config_manager: ConfigurationManager instance

    EXAMPLE:
        cm = ConfigurationManager()
        formatter = ConfigExcelFormatter(cm)
        formatter.export_program("P4M", "outputs/p4m_configs.xlsx")

    FORMATTING FEATURES:
    - Single "Configuration Matrix" sheet (replaces multiple sheets)
    - Category grouping with section headers
    - Freeze panes (config names stay visible when scrolling)
    - Yellow highlighting for overrides
    - "—" (em dash) for inherited values
    - Professional color scheme
    """

    # =========================================================================
    # STYLE DEFINITIONS - Professional color scheme
    # =========================================================================

    # Primary header style (dark blue with white text)
    HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Alternating row colors for readability
    ALT_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # Override highlight (amber/gold for attention)
    OVERRIDE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    OVERRIDE_FONT = Font(bold=True)

    # Category/section headers
    CATEGORY_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    CATEGORY_FONT = Font(bold=True, size=11)

    # Status colors
    ACTIVE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    INACTIVE_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Title styling
    TITLE_FONT = Font(size=18, bold=True, color="2F5496")
    SUBTITLE_FONT = Font(size=12, bold=True, color="595959")

    # Borders
    THIN_BORDER = Border(
        left=Side(style='thin', color='B4B4B4'),
        right=Side(style='thin', color='B4B4B4'),
        top=Side(style='thin', color='B4B4B4'),
        bottom=Side(style='thin', color='B4B4B4')
    )
    HEADER_BORDER = Border(
        left=Side(style='thin', color='1F4E79'),
        right=Side(style='thin', color='1F4E79'),
        top=Side(style='medium', color='1F4E79'),
        bottom=Side(style='medium', color='1F4E79')
    )

    # Default cell alignment
    DEFAULT_ALIGNMENT = Alignment(vertical='center', wrap_text=True)
    CENTER_ALIGNMENT = Alignment(horizontal='center', vertical='center')

    def __init__(self, config_manager: ConfigurationManager):
        """Initialize with a ConfigurationManager."""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl required. Install with: pip install openpyxl")

        self.cm = config_manager
        self.conn = config_manager.conn

    def export_program(self, program_prefix: str, output_path: str,
                       include_audit: bool = True,
                       include_providers: bool = True) -> str:
        """
        Export all configurations for a program to Excel using Configuration Matrix view.

        PURPOSE: Generate a single-sheet matrix showing all configs at all levels
                 Replaces the old multi-sheet approach with one clean view

        PARAMETERS:
            program_prefix: Program prefix (e.g., "P4M")
            output_path: Path for output file
            include_audit: Whether to include audit history sheet
            include_providers: Whether to include providers sheet

        RETURNS:
            str: Path to generated file

        MATRIX FORMAT:
            Columns: Config Key | Display Name | Program Default | Clinic | Loc1 | Loc2 | ...
            Rows: Grouped by category with section headers
            Values: "—" for inherited, actual value if set at that level
            Styling: Yellow highlight for overrides (differs from parent)
        """
        # Get program
        program = self.cm.get_program_by_prefix(program_prefix)
        if not program:
            raise ValueError(f"Program not found: {program_prefix}")

        program_id = program['program_id']

        # Create workbook
        wb = Workbook()

        # Create the Configuration Matrix (main view)
        self._create_configuration_matrix_sheet(wb, program)

        # Optional additional sheets
        if include_providers:
            self._create_providers_sheet(wb, program_id)

        if include_audit:
            self._create_audit_sheet(wb, program_id)

        # Remove default sheet if others exist
        if 'Sheet' in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb['Sheet']

        # Save
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))

        print(f"Exported Configuration Matrix to: {output_path}")
        return str(output_path)

    def _create_configuration_matrix_sheet(self, wb: Workbook, program: Dict) -> None:
        """
        Create the Configuration Matrix sheet - the primary view.

        PURPOSE: Single sheet showing all configs across all hierarchy levels
                 Replaces All Configs, Overrides, and By Category sheets

        LAYOUT:
            Row 1: Title row
            Row 2: Headers (Config Key | Display Name | Program Default | Location1 | Location2...)
            Row 3+: Data rows grouped by category

        COLUMNS:
            A: Config Key (technical name)
            B: Display Name (human-readable)
            C: Program Default value
            D+: Location values (or "—" if inherited)

        NOTE: No Clinic column - "Portland" etc. are geographic AREAS, not clinics.
        The actual clinics ARE the service locations (PCI Breast Surgery West, etc.)
        Hierarchy is: Program → Locations (no intermediate "area" level in export)

        FORMATTING:
            - Category headers span all columns (blue background)
            - Yellow fill for overrides (value differs from parent)
            - "—" (em dash) for inherited values
            - Freeze panes at row 2, column C
        """
        ws = wb.create_sheet("Configuration Matrix", 0)

        program_id = program['program_id']

        # =====================================================================
        # STEP 1: Get hierarchy (clinic and locations)
        # =====================================================================
        # NOTE: In the database, we still have clinic records, but in the EXPORT
        # we skip the clinic level because "Portland" etc. are geographic areas,
        # not actual service locations (clinics). The locations ARE the clinics.
        hierarchy = self.cm.get_program_hierarchy(program_id)
        clinics = hierarchy.get('clinics', [])

        # Collect ALL locations from all "clinics" (areas)
        # The clinic_id is still needed for database queries, but not shown in export
        all_locations = []
        clinic_id = None  # We'll use the first clinic's ID for queries if needed

        for clinic in clinics:
            if clinic_id is None:
                clinic_id = clinic['clinic_id']
            for loc in clinic.get('locations', []):
                all_locations.append({
                    'location_id': loc['location_id'],
                    'name': loc['name'],
                    'clinic_id': clinic['clinic_id']  # Keep track for queries
                })

        # =====================================================================
        # STEP 2: Build column headers
        # =====================================================================
        # Columns: Config Key | Display Name | Program Default | Loc1 | Loc2 | ...
        # NOTE: No clinic/area column - locations ARE the clinics
        headers = ["Config Key", "Display Name", "Program Default"]

        for loc in all_locations:
            headers.append(loc['name'])

        num_cols = len(headers)

        # =====================================================================
        # STEP 3: Write title row
        # =====================================================================
        ws['A1'] = f"Configuration Matrix: {program['name']}"
        ws['A1'].font = self.TITLE_FONT
        ws.merge_cells(f'A1:{get_column_letter(num_cols)}1')
        ws.row_dimensions[1].height = 30

        # =====================================================================
        # STEP 4: Write column headers (row 2)
        # =====================================================================
        self._apply_header_style(ws, headers, row=2)

        # =====================================================================
        # STEP 5: Get configs grouped by category
        # =====================================================================
        cursor = self.conn.cursor()

        # Get all categories in order
        # Note: Categories are defined inline in config_definitions, not in a separate table
        cursor.execute("""
            SELECT DISTINCT category
            FROM config_definitions
            ORDER BY display_order, category
        """)
        raw_categories = cursor.fetchall()

        # Transform to expected format with display names
        categories = []
        for row in raw_categories:
            cat = row['category']
            # Generate display name from category key (e.g., 'helpdesk' -> 'HELPDESK')
            display = cat.upper().replace('_', ' ')
            categories.append({
                'category': cat,
                'category_display': display
            })

        # =====================================================================
        # STEP 6: Write data rows grouped by category
        # =====================================================================
        row = 3

        for cat in categories:
            category = cat['category']
            category_display = cat['category_display'] or category.upper().replace('_', ' ')

            # Category header row (spans all columns)
            ws.cell(row=row, column=1, value=category_display)
            ws.cell(row=row, column=1).font = self.CATEGORY_FONT
            ws.cell(row=row, column=1).fill = self.CATEGORY_FILL
            ws.merge_cells(f'A{row}:{get_column_letter(num_cols)}{row}')
            row += 1

            # Get configs in this category
            cursor.execute("""
                SELECT * FROM config_definitions
                WHERE category = ?
                ORDER BY display_order, config_key
            """, (category,))
            definitions = cursor.fetchall()

            for defn in definitions:
                config_key = defn['config_key']
                display_name = defn['display_name']

                # Write config key and display name
                ws.cell(row=row, column=1, value=config_key)
                ws.cell(row=row, column=2, value=display_name)

                # Get value at PROGRAM level
                program_config = self.cm.get_config(config_key, program_id)
                program_value = program_config['value']

                # Write program default
                ws.cell(row=row, column=3, value=program_value or "—")
                # Program level is never "inherited" - it's the base

                # Get values at each LOCATION level
                # NOTE: Locations start at column 4 (D) - no clinic column
                col = 4
                for loc in all_locations:
                    loc_id = loc['location_id']
                    loc_clinic_id = loc['clinic_id']

                    # Query using the location's actual clinic_id
                    location_config = self.cm.get_config(config_key, program_id, loc_clinic_id, loc_id)
                    location_value = location_config['value']
                    location_level = location_config['effective_level']
                    location_is_override = location_config['is_override']

                    if location_level == 'location':
                        # Value is SET at location level - show it
                        cell_loc = ws.cell(row=row, column=col, value=location_value or "—")
                        # Highlight if it's an override
                        if location_is_override:
                            cell_loc.fill = self.OVERRIDE_FILL
                            cell_loc.font = self.OVERRIDE_FONT
                    else:
                        # Value is INHERITED from program - show "—"
                        ws.cell(row=row, column=col, value="—")

                    col += 1

                row += 1

            # Blank row after category (optional, for visual grouping)
            row += 1

        # =====================================================================
        # STEP 7: Apply formatting
        # =====================================================================
        # Freeze panes: freeze row 2 and columns A-B
        ws.freeze_panes = 'C3'

        # Auto-size columns
        ws.column_dimensions['A'].width = 25  # Config Key
        ws.column_dimensions['B'].width = 30  # Display Name
        ws.column_dimensions['C'].width = 20  # Program Default

        # Location columns start at D (column 4) - no clinic column
        for i, loc in enumerate(all_locations):
            col_letter = get_column_letter(4 + i)
            # Use wider columns for location names
            ws.column_dimensions[col_letter].width = 22

        # Apply borders to data cells
        for r in range(2, row):
            for c in range(1, num_cols + 1):
                cell = ws.cell(row=r, column=c)
                if cell.border == Border():  # Only if no border yet
                    cell.border = self.THIN_BORDER
                if cell.alignment == Alignment():
                    cell.alignment = self.DEFAULT_ALIGNMENT

        # Add generation timestamp at bottom
        ws.cell(row=row + 1, column=1,
                value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        ws.cell(row=row + 1, column=1).font = Font(italic=True, color="808080", size=9)

    def export_clinic(self, program_prefix: str, clinic_name: str,
                      output_path: str) -> str:
        """
        Export configurations for a specific clinic.

        PURPOSE: Generate clinic-specific configuration report
        """
        program = self.cm.get_program_by_prefix(program_prefix)
        if not program:
            raise ValueError(f"Program not found: {program_prefix}")

        clinic = self.cm.get_clinic_by_name(program['program_id'], clinic_name)
        if not clinic:
            raise ValueError(f"Clinic not found: {clinic_name}")

        wb = Workbook()

        self._create_clinic_summary_sheet(wb, program, clinic)
        self._create_clinic_configs_sheet(wb, program['program_id'], clinic['clinic_id'])
        self._create_clinic_locations_sheet(wb, clinic['clinic_id'])
        self._create_providers_sheet(wb, program['program_id'], clinic['clinic_id'])

        if 'Sheet' in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb['Sheet']

        wb.save(str(output_path))
        print(f"Exported clinic configurations to: {output_path}")
        return str(output_path)

    def _create_providers_sheet(self, wb: Workbook, program_id: str,
                                 clinic_id: str = None) -> None:
        """Create sheet with provider roster."""
        ws = wb.create_sheet("Providers")

        headers = ["Location", "Provider Name", "NPI", "Role", "Status"]

        providers = self.cm.get_providers(clinic_id=clinic_id) if clinic_id else []

        if not providers:
            # Get all providers for program
            cursor = self.conn.cursor()
            cursor.execute("""
                SELECT p.*, l.name as location_name
                FROM providers p
                JOIN locations l ON p.location_id = l.location_id
                JOIN clinics c ON l.clinic_id = c.clinic_id
                WHERE c.program_id = ?
                ORDER BY l.name, p.name
            """, (program_id,))
            providers = cursor.fetchall()

        row = 2
        for prov in providers:
            ws.cell(row=row, column=1, value=prov['location_name'])
            ws.cell(row=row, column=2, value=prov['name'])
            ws.cell(row=row, column=3, value=prov['npi'])
            ws.cell(row=row, column=4, value=prov['role'])

            # Status with conditional formatting
            status = "Active" if prov['is_active'] else "Inactive"
            status_cell = ws.cell(row=row, column=5, value=status)
            self._apply_status_formatting(status_cell, status)

            row += 1

        if row == 2:
            ws.cell(row=2, column=1, value="No providers found")
            ws.cell(row=2, column=1).font = Font(italic=True, color="808080")

        # Apply comprehensive formatting
        self._format_data_sheet(ws, headers)

    def _create_audit_sheet(self, wb: Workbook, program_id: str) -> None:
        """Create sheet with recent audit history."""
        ws = wb.create_sheet("Audit History")

        headers = ["Date", "Config Key", "Old Value", "New Value",
                   "Changed By", "Reason", "Source Doc"]

        cursor = self.conn.cursor()
        cursor.execute("""
            SELECT * FROM config_history
            WHERE program_id = ?
            ORDER BY changed_date DESC
            LIMIT 100
        """, (program_id,))

        row = 2
        for entry in cursor.fetchall():
            # Format date nicely if possible
            date_val = entry['changed_date']
            if date_val:
                try:
                    # Try to parse and reformat
                    dt = datetime.fromisoformat(date_val.replace('Z', '+00:00'))
                    date_val = dt.strftime("%Y-%m-%d %H:%M")
                except (ValueError, AttributeError):
                    pass

            ws.cell(row=row, column=1, value=date_val)
            ws.cell(row=row, column=2, value=entry['config_key'])
            ws.cell(row=row, column=3, value=entry['old_value'])
            ws.cell(row=row, column=4, value=entry['new_value'])
            ws.cell(row=row, column=5, value=entry['changed_by'])
            ws.cell(row=row, column=6, value=entry['change_reason'])
            ws.cell(row=row, column=7, value=entry['source_document'])

            # Highlight new entries (no old value = new record)
            if not entry['old_value']:
                for col in range(1, 8):
                    ws.cell(row=row, column=col).fill = self.ACTIVE_FILL

            row += 1

        if row == 2:
            ws.cell(row=2, column=1, value="No audit history found")
            ws.cell(row=2, column=1).font = Font(italic=True, color="808080")

        # Apply comprehensive formatting
        self._format_data_sheet(ws, headers)

    def _create_clinic_summary_sheet(self, wb: Workbook, program: Dict,
                                      clinic: Dict) -> None:
        """Create summary sheet for a clinic."""
        ws = wb.create_sheet("Summary", 0)

        ws['A1'] = f"Clinic Configuration: {clinic['name']}"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')

        info_data = [
            ("Program:", program['name']),
            ("Clinic:", clinic['name']),
            ("Clinic ID:", clinic['clinic_id']),
            ("Status:", clinic.get('status', 'Active')),
            ("Generated:", datetime.now().strftime("%Y-%m-%d %H:%M"))
        ]

        row = 3
        for label, value in info_data:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1

        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 40

    def _create_clinic_configs_sheet(self, wb: Workbook, program_id: str,
                                      clinic_id: str) -> None:
        """Create configs sheet for a specific clinic."""
        ws = wb.create_sheet("Configurations")

        headers = ["Config Key", "Display Name", "Category",
                   "Effective Value", "Level", "Is Override"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT

        cursor = self.conn.cursor()
        cursor.execute("""
            SELECT * FROM config_definitions ORDER BY category, display_order
        """)

        row = 2
        for defn in cursor.fetchall():
            config = self.cm.get_config(defn['config_key'], program_id, clinic_id)

            ws.cell(row=row, column=1, value=defn['config_key'])
            ws.cell(row=row, column=2, value=defn['display_name'])
            ws.cell(row=row, column=3, value=defn['category'])
            ws.cell(row=row, column=4, value=config['value'])
            ws.cell(row=row, column=5, value=config['effective_level'])
            ws.cell(row=row, column=6, value="Yes" if config['is_override'] else "No")

            if config['is_override']:
                for col in range(1, 7):
                    ws.cell(row=row, column=col).fill = self.OVERRIDE_FILL

            row += 1

        self._auto_size_columns(ws, headers)

    def _create_clinic_locations_sheet(self, wb: Workbook, clinic_id: str) -> None:
        """Create sheet with locations under a clinic."""
        ws = wb.create_sheet("Locations")

        headers = ["Location Name", "Code", "Status", "Provider Count"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT

        cursor = self.conn.cursor()
        cursor.execute("""
            SELECT l.*,
                   (SELECT COUNT(*) FROM providers p
                    WHERE p.location_id = l.location_id AND p.is_active = TRUE) as provider_count
            FROM locations l
            WHERE l.clinic_id = ?
            ORDER BY l.name
        """, (clinic_id,))

        row = 2
        for loc in cursor.fetchall():
            ws.cell(row=row, column=1, value=loc['name'])
            ws.cell(row=row, column=2, value=loc['code'])
            ws.cell(row=row, column=3, value=loc['status'])
            ws.cell(row=row, column=4, value=loc['provider_count'])
            row += 1

        self._auto_size_columns(ws, headers)

    def _auto_size_columns(self, ws, headers: List[str],
                           min_width: int = 10, max_width: int = 50) -> None:
        """
        Auto-size columns based on content with min/max bounds.

        PARAMETERS:
            ws: Worksheet to adjust
            headers: List of header names
            min_width: Minimum column width
            max_width: Maximum column width (prevents overly wide columns)
        """
        for i, header in enumerate(headers, 1):
            max_length = len(header)

            # Sample first 100 rows for performance
            for row in range(2, min(ws.max_row + 1, 102)):
                cell = ws.cell(row=row, column=i)
                if cell.value:
                    # Handle multi-line values
                    cell_len = max(len(line) for line in str(cell.value).split('\n'))
                    max_length = max(max_length, cell_len)

            # Apply bounds
            adjusted_width = max(min_width, min(max_length + 2, max_width))
            ws.column_dimensions[get_column_letter(i)].width = adjusted_width

    def _apply_header_style(self, ws, headers: List[str], row: int = 1) -> None:
        """
        Apply consistent header styling to a row.

        PARAMETERS:
            ws: Worksheet
            headers: List of header names
            row: Row number (default 1)
        """
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.HEADER_BORDER

        # Set row height for header
        ws.row_dimensions[row].height = 25

    def _apply_alternating_rows(self, ws, start_row: int = 2,
                                  end_row: int = None, num_cols: int = None) -> None:
        """
        Apply alternating row colors for readability.

        PARAMETERS:
            ws: Worksheet
            start_row: First data row (after headers)
            end_row: Last row (defaults to max_row)
            num_cols: Number of columns to style
        """
        if end_row is None:
            end_row = ws.max_row
        if num_cols is None:
            num_cols = ws.max_column

        for row in range(start_row, end_row + 1):
            fill = self.ALT_ROW_FILL if row % 2 == 0 else self.WHITE_FILL
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=row, column=col)
                # Don't override special fills (like overrides)
                if cell.fill.start_color.index == '00000000' or \
                   cell.fill.start_color.index == 'FFFFFF':
                    cell.fill = fill
                cell.border = self.THIN_BORDER
                cell.alignment = self.DEFAULT_ALIGNMENT

    def _add_freeze_panes(self, ws, freeze_cell: str = 'A2') -> None:
        """
        Freeze panes so headers stay visible when scrolling.

        PARAMETERS:
            ws: Worksheet
            freeze_cell: Cell reference for freeze point (default A2 = freeze row 1)
        """
        ws.freeze_panes = freeze_cell

    def _add_auto_filter(self, ws, start_col: int = 1, end_col: int = None,
                          header_row: int = 1) -> None:
        """
        Add auto-filter to header row for easy sorting/filtering.

        PARAMETERS:
            ws: Worksheet
            start_col: First column for filter
            end_col: Last column for filter (defaults to max)
            header_row: Row containing headers
        """
        if end_col is None:
            end_col = ws.max_column

        if ws.max_row > 1:  # Only add filter if there's data
            start_cell = f"{get_column_letter(start_col)}{header_row}"
            end_cell = f"{get_column_letter(end_col)}{ws.max_row}"
            ws.auto_filter.ref = f"{start_cell}:{end_cell}"

    def _format_data_sheet(self, ws, headers: List[str], has_overrides: bool = False,
                           override_col: int = None) -> None:
        """
        Apply complete formatting to a data sheet.

        Combines header styling, alternating rows, freeze panes, and auto-filter.

        PARAMETERS:
            ws: Worksheet to format
            headers: List of column headers
            has_overrides: If True, applies override highlighting
            override_col: Column index that indicates if row is override
        """
        # Apply header styling
        self._apply_header_style(ws, headers)

        # Freeze header row
        self._add_freeze_panes(ws, 'A2')

        # Auto-size columns
        self._auto_size_columns(ws, headers)

        # Apply alternating row colors
        self._apply_alternating_rows(ws, start_row=2, num_cols=len(headers))

        # Add auto-filter
        self._add_auto_filter(ws)

    def _apply_status_formatting(self, cell, status: str) -> None:
        """
        Apply conditional formatting based on status value.

        PARAMETERS:
            cell: Cell to format
            status: Status value ('Active', 'Inactive', etc.)
        """
        status_lower = str(status).lower() if status else ''

        if status_lower in ('active', 'yes', 'true', 'enabled'):
            cell.fill = self.ACTIVE_FILL
        elif status_lower in ('inactive', 'no', 'false', 'disabled'):
            cell.fill = self.INACTIVE_FILL


# ============================================================================
# MODULE TEST
# ============================================================================

if __name__ == "__main__":
    print("Testing ConfigExcelFormatter...")

    cm = ConfigurationManager()
    cm.initialize_schema()
    cm.load_definitions_from_yaml()

    # Create test data
    program_id = cm.create_program("Excel Export Test", "EXT")
    clinic_id = cm.create_clinic(program_id, "Test Clinic", "TCLI")
    location_id = cm.create_location(clinic_id, "Test Location")
    cm.add_provider(location_id, "Test Provider, MD", npi="1234567890")
    cm.set_config('helpdesk_phone', '555-1234', program_id, clinic_id)

    # Export
    formatter = ConfigExcelFormatter(cm)
    output_path = "outputs/test_export.xlsx"
    formatter.export_program("EXT", output_path)

    print(f"\nExported to: {output_path}")
    print("Tests complete!")
