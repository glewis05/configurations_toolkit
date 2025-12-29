"""
Quick Update Manager - Simplified interface for common updates

PURPOSE: Provide easy-to-use methods for frequent update operations
         without needing to know internal database structure

R EQUIVALENT: Like a wrapper function package (e.g., dplyr) that
provides simple verbs for common data operations

AVIATION ANALOGY: Like a Quick Reference Handbook - simplified
procedures for common operations without flipping through manuals

AUTHOR: Glen Lewis
DATE: 2024
"""

import json
import re
from pathlib import Path
from typing import Dict, List, Optional, Any
from datetime import datetime

import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from database.config_manager import ConfigurationManager


class QuickUpdateManager:
    """
    PURPOSE: Simplified interface for common update operations

    Makes it easy to:
    - Update provider NPIs
    - Change test codes
    - Update phone numbers
    - Bulk update locations

    PARAMETERS:
        config_manager: ConfigurationManager instance

    EXAMPLE:
        cm = ConfigurationManager()
        qm = QuickUpdateManager(cm)
        qm.update_provider_npi("Christine Kemp", "1215158639")
    """

    def __init__(self, config_manager: ConfigurationManager):
        """
        Initialize with a ConfigurationManager.

        WHY THIS APPROACH: We delegate to ConfigurationManager for
        actual operations but provide simpler method signatures.
        """
        self.cm = config_manager
        self.conn = config_manager.conn

    # ========================================================================
    # PROVIDER UPDATES
    # ========================================================================

    def update_provider_npi(self, provider_name: str, new_npi: str,
                            location_id: str = None,
                            program_prefix: str = None) -> int:
        """
        Update NPI for a provider by name.

        PURPOSE: Quickly update an NPI when it changes or was incorrect

        PARAMETERS:
            provider_name: Full or partial provider name
            new_npi: New NPI number (10 digits)
            location_id: Optional - if provided, only update at this location
            program_prefix: Optional - limit search to this program

        RETURNS:
            int: Number of providers updated

        EXAMPLE:
            qm.update_provider_npi("Christine Kemp", "1215158639")
            # Updates all providers named Christine Kemp

            qm.update_provider_npi("Kemp", "1215158639", program_prefix="P4M")
            # Updates providers with Kemp in name, in P4M program only
        """
        cursor = self.conn.cursor()

        # Validate NPI format (10 digits)
        npi_clean = re.sub(r'\D', '', new_npi)
        if len(npi_clean) != 10:
            raise ValueError(f"NPI must be 10 digits, got: {new_npi}")

        # Build query based on parameters
        if location_id:
            # Specific location
            cursor.execute("""
                SELECT provider_id, name, npi, location_id
                FROM providers
                WHERE location_id = ? AND name LIKE ?
            """, (location_id, f"%{provider_name}%"))
        elif program_prefix:
            # All locations in a program
            cursor.execute("""
                SELECT p.provider_id, p.name, p.npi, p.location_id
                FROM providers p
                JOIN locations l ON p.location_id = l.location_id
                JOIN clinics c ON l.clinic_id = c.clinic_id
                JOIN programs pr ON c.program_id = pr.program_id
                WHERE pr.prefix = ? AND p.name LIKE ?
            """, (program_prefix, f"%{provider_name}%"))
        else:
            # All matching providers
            cursor.execute("""
                SELECT provider_id, name, npi, location_id
                FROM providers WHERE name LIKE ?
            """, (f"%{provider_name}%",))

        providers = cursor.fetchall()

        if not providers:
            print(f"No providers found matching: {provider_name}")
            return 0

        count = 0
        for prov in providers:
            old_npi = prov['npi']
            self.cm.update_provider(prov['provider_id'], npi=npi_clean)
            print(f"Updated {prov['name']}: {old_npi} → {npi_clean}")
            count += 1

        return count

    def add_provider_to_location(self, location_name: str, provider_name: str,
                                  npi: str = None, role: str = 'Ordering Provider',
                                  clinic_name: str = None) -> int:
        """
        Add a new provider to a location by name lookup.

        PURPOSE: Add provider without needing to know location_id

        PARAMETERS:
            location_name: Full or partial location name
            provider_name: Provider's full name
            npi: Optional NPI number
            role: Provider role (default: Ordering Provider)
            clinic_name: Optional clinic name to narrow search

        RETURNS:
            int: provider_id of created provider
        """
        cursor = self.conn.cursor()

        # Find the location
        if clinic_name:
            cursor.execute("""
                SELECT l.location_id, l.name, c.name as clinic_name
                FROM locations l
                JOIN clinics c ON l.clinic_id = c.clinic_id
                WHERE l.name LIKE ? AND c.name LIKE ?
            """, (f"%{location_name}%", f"%{clinic_name}%"))
        else:
            cursor.execute("""
                SELECT l.location_id, l.name, c.name as clinic_name
                FROM locations l
                JOIN clinics c ON l.clinic_id = c.clinic_id
                WHERE l.name LIKE ?
            """, (f"%{location_name}%",))

        locations = cursor.fetchall()

        if not locations:
            raise ValueError(f"No location found matching: {location_name}")
        if len(locations) > 1:
            loc_list = [f"{l['clinic_name']} / {l['name']}" for l in locations]
            raise ValueError(f"Multiple locations found. Specify clinic_name. Found: {loc_list}")

        location = locations[0]
        provider_id = self.cm.add_provider(
            location_id=location['location_id'],
            name=provider_name,
            npi=npi,
            role=role
        )

        print(f"Added provider {provider_name} to {location['name']}")
        return provider_id

    # ========================================================================
    # TEST CODE UPDATES
    # ========================================================================

    def update_test_code(self, clinic_name: str, new_code: str,
                         new_name: str = None, modifications: str = None,
                         program_prefix: str = None) -> bool:
        """
        Update default test code for a clinic.

        PURPOSE: Quickly change lab test when test panel changes

        PARAMETERS:
            clinic_name: Full or partial clinic name
            new_code: New test code
            new_name: Optional new test name
            modifications: Optional gene modifications text
            program_prefix: Optional program prefix to narrow search

        RETURNS:
            bool: True if updated successfully

        EXAMPLE:
            qm.update_test_code("Portland", "CAP123", new_name="Custom Breast Panel")
        """
        cursor = self.conn.cursor()

        # Find the clinic
        if program_prefix:
            cursor.execute("""
                SELECT c.clinic_id, c.name, p.program_id
                FROM clinics c
                JOIN programs p ON c.program_id = p.program_id
                WHERE c.name LIKE ? AND p.prefix = ?
            """, (f"%{clinic_name}%", program_prefix))
        else:
            cursor.execute("""
                SELECT c.clinic_id, c.name, c.program_id
                FROM clinics c
                WHERE c.name LIKE ?
            """, (f"%{clinic_name}%",))

        clinics = cursor.fetchall()

        if not clinics:
            raise ValueError(f"No clinic found matching: {clinic_name}")
        if len(clinics) > 1:
            clinic_list = [c['name'] for c in clinics]
            raise ValueError(f"Multiple clinics found. Specify program_prefix. Found: {clinic_list}")

        clinic = clinics[0]

        # Update the test code config
        self.cm.set_config(
            'lab_default_test_code',
            new_code,
            clinic['program_id'],
            clinic['clinic_id'],
            source='manual',
            rationale=f'Test code updated via QuickUpdateManager'
        )

        if new_name:
            self.cm.set_config(
                'lab_default_test_name',
                new_name,
                clinic['program_id'],
                clinic['clinic_id'],
                source='manual'
            )

        if modifications:
            self.cm.set_config(
                'lab_test_modifications',
                modifications,
                clinic['program_id'],
                clinic['clinic_id'],
                source='manual'
            )

        print(f"Updated test code for {clinic['name']}: {new_code}")
        return True

    # ========================================================================
    # PHONE NUMBER UPDATES
    # ========================================================================

    def update_phone(self, location_name: str, new_phone: str,
                     clinic_name: str = None) -> bool:
        """
        Update helpdesk phone for a location.

        PURPOSE: Quickly change phone number when it changes

        PARAMETERS:
            location_name: Full or partial location name
            new_phone: New phone number (any format)
            clinic_name: Optional clinic name to narrow search

        RETURNS:
            bool: True if updated successfully
        """
        cursor = self.conn.cursor()

        # Find the location
        if clinic_name:
            cursor.execute("""
                SELECT l.location_id, l.name, c.clinic_id, c.name as clinic_name, c.program_id
                FROM locations l
                JOIN clinics c ON l.clinic_id = c.clinic_id
                WHERE l.name LIKE ? AND c.name LIKE ?
            """, (f"%{location_name}%", f"%{clinic_name}%"))
        else:
            cursor.execute("""
                SELECT l.location_id, l.name, c.clinic_id, c.name as clinic_name, c.program_id
                FROM locations l
                JOIN clinics c ON l.clinic_id = c.clinic_id
                WHERE l.name LIKE ?
            """, (f"%{location_name}%",))

        locations = cursor.fetchall()

        if not locations:
            raise ValueError(f"No location found matching: {location_name}")
        if len(locations) > 1:
            loc_list = [f"{l['clinic_name']} / {l['name']}" for l in locations]
            raise ValueError(f"Multiple locations found. Specify clinic_name. Found: {loc_list}")

        location = locations[0]

        self.cm.set_config(
            'helpdesk_phone',
            new_phone,
            location['program_id'],
            location['clinic_id'],
            location['location_id'],
            source='manual',
            rationale='Phone updated via QuickUpdateManager'
        )

        print(f"Updated phone for {location['name']}: {new_phone}")
        return True

    # ========================================================================
    # HOURS UPDATES
    # ========================================================================

    def update_hours(self, location_name: str, open_time: str, close_time: str,
                     clinic_name: str = None) -> bool:
        """
        Update operating hours for a location.

        PURPOSE: Quickly change hours when schedule changes

        PARAMETERS:
            location_name: Full or partial location name
            open_time: Opening time (e.g., "08:00", "8:00 AM")
            close_time: Closing time (e.g., "17:00", "5:00 PM")
            clinic_name: Optional clinic name to narrow search

        RETURNS:
            bool: True if updated successfully
        """
        cursor = self.conn.cursor()

        # Normalize time format
        open_normalized = self._normalize_time(open_time)
        close_normalized = self._normalize_time(close_time)

        # Find the location
        if clinic_name:
            cursor.execute("""
                SELECT l.location_id, l.name, c.clinic_id, c.name as clinic_name, c.program_id
                FROM locations l
                JOIN clinics c ON l.clinic_id = c.clinic_id
                WHERE l.name LIKE ? AND c.name LIKE ?
            """, (f"%{location_name}%", f"%{clinic_name}%"))
        else:
            cursor.execute("""
                SELECT l.location_id, l.name, c.clinic_id, c.name as clinic_name, c.program_id
                FROM locations l
                JOIN clinics c ON l.clinic_id = c.clinic_id
                WHERE l.name LIKE ?
            """, (f"%{location_name}%",))

        locations = cursor.fetchall()

        if not locations:
            raise ValueError(f"No location found matching: {location_name}")
        if len(locations) > 1:
            loc_list = [f"{l['clinic_name']} / {l['name']}" for l in locations]
            raise ValueError(f"Multiple locations found. Specify clinic_name. Found: {loc_list}")

        location = locations[0]

        self.cm.set_config(
            'hours_open',
            open_normalized,
            location['program_id'],
            location['clinic_id'],
            location['location_id'],
            source='manual'
        )

        self.cm.set_config(
            'hours_close',
            close_normalized,
            location['program_id'],
            location['clinic_id'],
            location['location_id'],
            source='manual'
        )

        print(f"Updated hours for {location['name']}: {open_normalized} - {close_normalized}")
        return True

    def _normalize_time(self, time_str: str) -> str:
        """
        Normalize time string to 24-hour format.

        EXAMPLES:
            "8:00 AM" → "08:00"
            "5:00 PM" → "17:00"
            "08:00" → "08:00"
        """
        # Already in 24-hour format
        if re.match(r'^\d{1,2}:\d{2}$', time_str.strip()):
            parts = time_str.strip().split(':')
            return f"{int(parts[0]):02d}:{parts[1]}"

        # 12-hour format with AM/PM
        match = re.match(r'(\d{1,2}):(\d{2})\s*(AM|PM)', time_str.strip(), re.IGNORECASE)
        if match:
            hour = int(match.group(1))
            minute = match.group(2)
            period = match.group(3).upper()

            if period == 'PM' and hour != 12:
                hour += 12
            elif period == 'AM' and hour == 12:
                hour = 0

            return f"{hour:02d}:{minute}"

        # Return as-is if can't parse
        return time_str

    # ========================================================================
    # BULK UPDATES
    # ========================================================================

    def bulk_update_from_excel(self, excel_path: str, dry_run: bool = True) -> Dict:
        """
        Update multiple configs from Excel file.

        PURPOSE: Apply batch updates from a spreadsheet

        PARAMETERS:
            excel_path: Path to Excel file with updates
            dry_run: If True, show what would change without applying

        Expected Excel columns:
        - Location (or Clinic): Name of entity to update
        - Config Key: The configuration key (e.g., 'helpdesk_phone')
        - New Value: The new value to set
        - Rationale: Optional reason for change

        RETURNS:
            Dict with counts: {'applied': N, 'skipped': N, 'errors': [...]}

        EXAMPLE:
            result = qm.bulk_update_from_excel("updates.xlsx", dry_run=False)
            print(f"Applied {result['applied']} updates")
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl required. Install with: pip install openpyxl")

        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

        # Get headers
        headers = [cell.value.lower() if cell.value else '' for cell in ws[1]]

        # Map columns
        col_map = {}
        for i, h in enumerate(headers):
            if 'location' in h:
                col_map['location'] = i
            elif 'clinic' in h:
                col_map['clinic'] = i
            elif 'config' in h or 'key' in h:
                col_map['config_key'] = i
            elif 'value' in h:
                col_map['value'] = i
            elif 'rationale' in h or 'reason' in h:
                col_map['rationale'] = i

        if 'config_key' not in col_map or 'value' not in col_map:
            raise ValueError("Excel must have 'Config Key' and 'Value' columns")

        result = {'applied': 0, 'skipped': 0, 'errors': []}

        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                config_key = row[col_map['config_key']]
                new_value = str(row[col_map['value']]) if row[col_map['value']] else None
                rationale = row[col_map.get('rationale', len(row))] if col_map.get('rationale') else None

                if not config_key or not new_value:
                    result['skipped'] += 1
                    continue

                location_name = row[col_map.get('location')] if col_map.get('location') else None
                clinic_name = row[col_map.get('clinic')] if col_map.get('clinic') else None

                if dry_run:
                    print(f"Would update {config_key} at {location_name or clinic_name}: {new_value}")
                    result['applied'] += 1
                else:
                    # Find the entity and update
                    if location_name:
                        self._apply_location_update(location_name, clinic_name,
                                                    config_key, new_value, rationale)
                    elif clinic_name:
                        self._apply_clinic_update(clinic_name, config_key,
                                                  new_value, rationale)
                    result['applied'] += 1

            except Exception as e:
                result['errors'].append({'row': row, 'error': str(e)})

        if dry_run:
            print(f"\nDRY RUN: Would apply {result['applied']} updates")
        else:
            print(f"Applied {result['applied']} updates, {len(result['errors'])} errors")

        return result

    def _apply_location_update(self, location_name: str, clinic_name: str,
                               config_key: str, value: str, rationale: str) -> None:
        """Apply a config update at location level."""
        cursor = self.conn.cursor()

        query = """
            SELECT l.location_id, c.clinic_id, c.program_id
            FROM locations l
            JOIN clinics c ON l.clinic_id = c.clinic_id
            WHERE l.name LIKE ?
        """
        params = [f"%{location_name}%"]

        if clinic_name:
            query += " AND c.name LIKE ?"
            params.append(f"%{clinic_name}%")

        cursor.execute(query, params)
        loc = cursor.fetchone()

        if not loc:
            raise ValueError(f"Location not found: {location_name}")

        self.cm.set_config(config_key, value, loc['program_id'],
                           loc['clinic_id'], loc['location_id'],
                           source='bulk_import', rationale=rationale)

    def _apply_clinic_update(self, clinic_name: str, config_key: str,
                             value: str, rationale: str) -> None:
        """Apply a config update at clinic level."""
        cursor = self.conn.cursor()

        cursor.execute("""
            SELECT clinic_id, program_id FROM clinics WHERE name LIKE ?
        """, (f"%{clinic_name}%",))
        clinic = cursor.fetchone()

        if not clinic:
            raise ValueError(f"Clinic not found: {clinic_name}")

        self.cm.set_config(config_key, value, clinic['program_id'],
                           clinic['clinic_id'],
                           source='bulk_import', rationale=rationale)

    # ========================================================================
    # UTILITY METHODS
    # ========================================================================

    def show_locations(self, program_prefix: str = None) -> List[Dict]:
        """
        Show all locations, optionally filtered by program.

        PURPOSE: Quick reference to see available locations
        """
        cursor = self.conn.cursor()

        if program_prefix:
            cursor.execute("""
                SELECT p.name as program_name, p.prefix,
                       c.name as clinic_name, l.name as location_name,
                       l.location_id
                FROM locations l
                JOIN clinics c ON l.clinic_id = c.clinic_id
                JOIN programs p ON c.program_id = p.program_id
                WHERE p.prefix = ?
                ORDER BY p.name, c.name, l.name
            """, (program_prefix,))
        else:
            cursor.execute("""
                SELECT p.name as program_name, p.prefix,
                       c.name as clinic_name, l.name as location_name,
                       l.location_id
                FROM locations l
                JOIN clinics c ON l.clinic_id = c.clinic_id
                JOIN programs p ON c.program_id = p.program_id
                ORDER BY p.name, c.name, l.name
            """)

        return [dict(row) for row in cursor.fetchall()]

    def show_providers(self, location_name: str = None,
                       program_prefix: str = None) -> List[Dict]:
        """
        Show all providers, optionally filtered.

        PURPOSE: Quick reference to see provider roster
        """
        cursor = self.conn.cursor()

        if location_name:
            cursor.execute("""
                SELECT pr.name, pr.npi, pr.role, l.name as location_name
                FROM providers pr
                JOIN locations l ON pr.location_id = l.location_id
                WHERE l.name LIKE ? AND pr.is_active = TRUE
                ORDER BY pr.name
            """, (f"%{location_name}%",))
        elif program_prefix:
            cursor.execute("""
                SELECT pr.name, pr.npi, pr.role, l.name as location_name
                FROM providers pr
                JOIN locations l ON pr.location_id = l.location_id
                JOIN clinics c ON l.clinic_id = c.clinic_id
                JOIN programs p ON c.program_id = p.program_id
                WHERE p.prefix = ? AND pr.is_active = TRUE
                ORDER BY l.name, pr.name
            """, (program_prefix,))
        else:
            cursor.execute("""
                SELECT pr.name, pr.npi, pr.role, l.name as location_name
                FROM providers pr
                JOIN locations l ON pr.location_id = l.location_id
                WHERE pr.is_active = TRUE
                ORDER BY l.name, pr.name
            """)

        return [dict(row) for row in cursor.fetchall()]


# ============================================================================
# MODULE TEST
# ============================================================================

if __name__ == "__main__":
    print("Testing QuickUpdateManager...")

    cm = ConfigurationManager()
    cm.initialize_schema()
    cm.load_definitions_from_yaml()

    qm = QuickUpdateManager(cm)

    # Create test data
    program_id = cm.create_program("Quick Update Test", "QUT")
    clinic_id = cm.create_clinic(program_id, "Portland Test Clinic", "PORT")
    location_id = cm.create_location(clinic_id, "Test Breast Surgery West", "LOC1")
    cm.add_provider(location_id, "Jane Smith, MD", npi="1234567890")

    # Test time normalization
    print("\nTime normalization tests:")
    print(f"'8:00 AM' → '{qm._normalize_time('8:00 AM')}'")
    print(f"'5:00 PM' → '{qm._normalize_time('5:00 PM')}'")
    print(f"'14:30' → '{qm._normalize_time('14:30')}'")

    # Test showing data
    print("\nLocations:")
    for loc in qm.show_locations("QUT"):
        print(f"  {loc['clinic_name']} / {loc['location_name']}")

    print("\nProviders:")
    for prov in qm.show_providers(program_prefix="QUT"):
        print(f"  {prov['name']} at {prov['location_name']}")

    print("\nTests complete!")
