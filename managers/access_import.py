"""
ACCESS IMPORT MODULE
====================
Import users, access grants, and training records from Excel files.

This module is useful when:
- Onboarding existing clinics with established user lists
- Migrating from another tracking system (spreadsheets, SharePoint, etc.)
- Bulk provisioning new users during clinic go-live
- Importing historical training records for compliance

Aviation Analogy:
    Think of this like migrating crew records from paper logbooks to a digital system.
    You have existing data about who's qualified for what aircraft (access grants),
    their medical certificates (training records), and basic crew info (users).
    This importer reads that data from Excel and loads it into the database,
    validating as it goes to catch issues early.

Column Matching:
    The importer uses flexible column matching - it looks for columns by multiple
    possible names. For example, "Name", "Full Name", and "User Name" all map to
    the name field. This makes it forgiving of different Excel formats.

Import Order:
    When importing from a multi-tab template, order matters:
    1. Users first (must exist before granting access)
    2. Access grants second (references users)
    3. Training records third (references users)
"""

from typing import Optional, Dict, List, Any, Tuple
from pathlib import Path
from datetime import datetime, date, timedelta
import re

# openpyxl for Excel reading
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from managers.access_manager import AccessManager


class AccessImporter:
    """
    PURPOSE: Import users, access grants, and training from Excel files

    This class provides methods to bulk-import data from Excel spreadsheets,
    handling the common task of onboarding clinics or migrating from other
    tracking systems.

    R EQUIVALENT:
        In R, you'd use readxl::read_excel() to read the data, then
        purrr::map() or lapply() to process each row. This class wraps
        those operations into methods like import_users(), import_access(), etc.

    ATTRIBUTES:
        am: AccessManager instance for database operations

    EXAMPLE:
        importer = AccessImporter(am)

        # Generate template for data entry
        importer.generate_import_template("Template.xlsx")

        # After filling in the template, import it
        results = importer.import_from_template("Template_Filled.xlsx")
        print(f"Imported {results['users']['imported']} users")

    COLUMN MATCHING:
        The importer uses flexible column matching. For example, these
        column headers all map to the "email" field:
        - "Email"
        - "Email Address"
        - "User Email"
        - "E-mail"

        This makes the importer forgiving of different Excel formats.
    """

    # =========================================================================
    # COLUMN MAPPINGS
    # =========================================================================
    # Define multiple possible column names for each field.
    # The importer will look for any of these names (case-insensitive).
    # First match wins, so put preferred names first.
    # =========================================================================

    # User import column mappings
    USER_COLUMNS = {
        'name': ['Name', 'Full Name', 'User Name', 'Employee Name', 'Staff Name'],
        'email': ['Email', 'Email Address', 'User Email', 'E-mail', 'E-Mail Address'],
        'organization': ['Organization', 'Company', 'Org', 'Employer', 'Department'],
        'is_business_associate': ['Is Business Associate', 'BA', 'External', 'Business Associate',
                                   'Is BA', 'External User', 'Contractor'],
        'status': ['Status', 'User Status', 'Account Status', 'Active'],
        'notes': ['Notes', 'Comments', 'Remarks', 'Description']
    }

    # Access import column mappings
    ACCESS_COLUMNS = {
        'user_email': ['User Email', 'Email', 'User', 'Email Address'],
        'program': ['Program', 'Program Name', 'Product', 'Application'],
        'clinic': ['Clinic', 'Clinic Name', 'Organization', 'Site'],
        'location': ['Location', 'Location Name', 'Facility', 'Department'],
        'role': ['Role', 'Access Level', 'Permission', 'Access Role', 'User Role'],
        'granted_date': ['Granted Date', 'Start Date', 'Effective Date', 'Access Date',
                         'Date Granted', 'Grant Date'],
        'granted_by': ['Granted By', 'Approved By', 'Manager', 'Approver', 'Requestor'],
        'reason': ['Reason', 'Justification', 'Notes', 'Purpose', 'Business Reason'],
        'ticket': ['Ticket', 'Ticket #', 'Reference', 'Request #', 'Ticket Number', 'SR#'],
        'review_cycle': ['Review Cycle', 'Review Frequency', 'Review Period', 'Recertification']
    }

    # Training import column mappings
    TRAINING_COLUMNS = {
        'user_email': ['User Email', 'Email', 'User', 'Email Address'],
        'training_type': ['Training Type', 'Training', 'Course', 'Course Name',
                          'Training Name', 'Module'],
        'responsibility': ['Training Responsibility', 'Responsibility', 'Maintained By',
                           'Record Owner', 'Training Owner'],
        'completed_date': ['Completed Date', 'Date Completed', 'Completion Date',
                           'Completed', 'Date', 'Training Date'],
        'expires_date': ['Expires Date', 'Expiration', 'Valid Until', 'Expiry Date',
                         'Expires', 'Expiration Date'],
        'certificate': ['Certificate', 'Certificate #', 'Reference', 'Cert #',
                        'Certificate Number', 'Certificate ID'],
        'status': ['Status', 'Training Status', 'Completion Status']
    }

    # =========================================================================
    # TRAINING TYPE DEFINITIONS
    # =========================================================================
    # Training types are categorized as Active (currently in use) or Reserved
    # (defined for future use but not yet implemented).
    #
    # Training Responsibility indicates who maintains the training records:
    #   - Client: The client organization tracks their own employee training
    #   - Propel Health: PHP maintains training records internally
    # =========================================================================

    # Active training types - currently in use
    ACTIVE_TRAINING_TYPES = {
        'HIPAA',                # Consolidated HIPAA Privacy & Security training
        'Cybersecurity',        # Annual cybersecurity awareness training
        'Application Training'  # Product-specific application training
    }

    # Reserved training types - defined for future use
    RESERVED_TRAINING_TYPES = {
        'SOC 2',    # SOC 2 compliance training (future)
        'HITRUST',  # HITRUST certification training (future)
        'Part 11'   # FDA 21 CFR Part 11 training (future)
    }

    # Combined set for validation - accepts both active and reserved types
    VALID_TRAINING_TYPES = ACTIVE_TRAINING_TYPES | RESERVED_TRAINING_TYPES

    # Valid training responsibility values
    VALID_TRAINING_RESPONSIBILITY = {
        'Client',       # Client organization maintains records
        'Propel Health' # PHP maintains records internally
    }

    # Style definitions for template generation
    HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    EXAMPLE_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    REQUIRED_FONT = Font(bold=True)

    def __init__(self, access_manager: AccessManager):
        """
        Initialize AccessImporter with an AccessManager instance.

        PURPOSE: Set up the importer with database access

        PARAMETERS:
            access_manager: AccessManager instance for database operations

        RAISES:
            ImportError: If openpyxl is not installed
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl required. Install with: pip install openpyxl")

        self.am = access_manager

    # =========================================================================
    # IMPORT METHODS
    # =========================================================================

    def import_users(
        self,
        excel_path: str,
        sheet_name: str = None,
        dry_run: bool = False
    ) -> Dict[str, Any]:
        """
        Import users from Excel.

        PURPOSE: Bulk import users from a spreadsheet

        R EQUIVALENT:
            users <- read_excel("users.xlsx") %>%
              mutate(user_id = map_chr(., create_user))

        PARAMETERS:
            excel_path: Path to Excel file
            sheet_name: Optional sheet name (defaults to first sheet)
            dry_run: If True, validate but don't actually import

        EXPECTED COLUMNS (flexible matching):
            - Name (required): Full name like "John Smith"
            - Email (required): Email address (must be unique)
            - Organization: Company name, defaults to 'Internal'
            - Is Business Associate: Yes/No/True/False, defaults to No
            - Status: Active/Inactive/Terminated, defaults to Active
            - Notes: Optional notes

        RETURNS:
            Dict with:
            - imported: Count of successfully imported users
            - skipped: Count of skipped rows (already exist)
            - errors: List of error messages with row numbers
            - user_ids: List of created user IDs

        AVIATION ANALOGY:
            Like importing crew member data from a roster spreadsheet.
            Each row is a crew member with their basic info. We validate
            that required fields are present and the email is unique.

        EXAMPLE:
            results = importer.import_users("new_hires.xlsx")
            print(f"Imported {results['imported']} users")
            for error in results['errors']:
                print(f"  Error: {error}")
        """
        wb = load_workbook(excel_path, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active

        # Find column mappings
        column_map = self._find_columns(ws, self.USER_COLUMNS)

        # Validate required columns
        if 'name' not in column_map:
            return {'imported': 0, 'skipped': 0, 'errors': ['Missing required column: Name'], 'user_ids': []}
        if 'email' not in column_map:
            return {'imported': 0, 'skipped': 0, 'errors': ['Missing required column: Email'], 'user_ids': []}

        results = {
            'imported': 0,
            'skipped': 0,
            'errors': [],
            'user_ids': []
        }

        # Process each row (skip header)
        for row_num in range(2, ws.max_row + 1):
            try:
                # Extract values using column map
                name = self._get_cell_value(ws, row_num, column_map.get('name'))
                email = self._get_cell_value(ws, row_num, column_map.get('email'))

                # Skip empty rows
                if not name and not email:
                    continue

                # Validate required fields
                if not name:
                    results['errors'].append(f"Row {row_num}: Missing name")
                    continue
                if not email:
                    results['errors'].append(f"Row {row_num}: Missing email")
                    continue

                # Validate email format
                if not self._is_valid_email(email):
                    results['errors'].append(f"Row {row_num}: Invalid email format: {email}")
                    continue

                # Check if user already exists
                existing = self.am.get_user(email=email)
                if existing:
                    results['skipped'] += 1
                    continue

                # Extract optional fields
                organization = self._get_cell_value(ws, row_num, column_map.get('organization')) or 'Internal'
                is_ba_raw = self._get_cell_value(ws, row_num, column_map.get('is_business_associate'))
                is_ba = self._parse_boolean(is_ba_raw)
                status = self._get_cell_value(ws, row_num, column_map.get('status')) or 'Active'
                notes = self._get_cell_value(ws, row_num, column_map.get('notes'))

                # Validate status
                if status not in ('Active', 'Inactive', 'Terminated'):
                    results['errors'].append(f"Row {row_num}: Invalid status: {status}")
                    continue

                # Create user (unless dry run)
                if not dry_run:
                    user_id = self.am.create_user(
                        name=name,
                        email=email,
                        organization=organization,
                        is_business_associate=is_ba,
                        notes=notes
                    )

                    # Update status if not Active
                    if status != 'Active':
                        self.am.update_user(user_id, changed_by='import', status=status)

                    results['user_ids'].append(user_id)

                results['imported'] += 1

            except Exception as e:
                results['errors'].append(f"Row {row_num}: {str(e)}")

        return results

    def import_access(
        self,
        excel_path: str,
        sheet_name: str = None,
        dry_run: bool = False
    ) -> Dict[str, Any]:
        """
        Import access grants from Excel.

        PURPOSE: Bulk import access grants from a spreadsheet

        R EQUIVALENT:
            access <- read_excel("access.xlsx") %>%
              left_join(users, by = "email") %>%
              mutate(access_id = pmap_int(., grant_access))

        PARAMETERS:
            excel_path: Path to Excel file
            sheet_name: Optional sheet name (defaults to first sheet)
            dry_run: If True, validate but don't actually import

        EXPECTED COLUMNS (flexible matching):
            - User Email (required): Email of existing user
            - Program (required): Program name or prefix
            - Clinic: Clinic name (optional, limits scope)
            - Location: Location name (optional, requires clinic)
            - Role (required): Read-Only, Read-Write, Read-Write-Order, Clinic-Manager,
                               Analytics-Only, Admin, or Auditor
            - Granted Date: When access was granted (defaults to today)
            - Granted By (required): Who approved the access
            - Reason: Business justification
            - Ticket: Reference to approval ticket
            - Review Cycle: Quarterly or Annual (defaults to Quarterly)

        RETURNS:
            Dict with:
            - imported: Count of successfully imported access grants
            - skipped: Count of skipped rows
            - conflicts: List of segregation of duties warnings
            - errors: List of error messages with row numbers

        PROCESS:
            1. Look up user by email (must exist)
            2. Look up program by name or prefix
            3. Look up clinic/location if provided
            4. Check for segregation of duties conflicts
            5. Create access grant
            6. Log to audit_history

        AVIATION ANALOGY:
            Like importing type rating records. Each row says which pilot
            (user) has which type rating (role) for which aircraft (program).
            We validate that the pilot exists and the aircraft is in our fleet.
        """
        wb = load_workbook(excel_path, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active

        # Find column mappings
        column_map = self._find_columns(ws, self.ACCESS_COLUMNS)

        # Validate required columns
        required = ['user_email', 'program', 'role', 'granted_by']
        missing = [col for col in required if col not in column_map]
        if missing:
            return {
                'imported': 0,
                'skipped': 0,
                'conflicts': [],
                'errors': [f"Missing required columns: {', '.join(missing)}"]
            }

        results = {
            'imported': 0,
            'skipped': 0,
            'conflicts': [],
            'errors': []
        }

        # Valid roles
        # - Read-Only: View only, single clinic scope
        # - Read-Write: View + Edit, single clinic scope
        # - Read-Write-Order: View + Edit + Order Tests, single clinic scope
        # - Clinic-Manager: View + Edit + Order Tests + Analytics, single clinic scope
        # - Analytics-Only: Aggregated analytics only, cross-clinic scope (no patient data)
        # - Admin: View + Edit + Analytics, cross-clinic scope (system-level)
        # - Auditor: Audit/compliance access
        valid_roles = {
            'Read-Only', 'Read-Write', 'Read-Write-Order', 'Clinic-Manager',
            'Analytics-Only', 'Admin', 'Auditor'
        }

        # Process each row (skip header)
        for row_num in range(2, ws.max_row + 1):
            try:
                # Extract values
                user_email = self._get_cell_value(ws, row_num, column_map.get('user_email'))
                program = self._get_cell_value(ws, row_num, column_map.get('program'))
                role = self._get_cell_value(ws, row_num, column_map.get('role'))
                granted_by = self._get_cell_value(ws, row_num, column_map.get('granted_by'))

                # Skip empty rows
                if not user_email and not program:
                    continue

                # Validate required fields
                if not user_email:
                    results['errors'].append(f"Row {row_num}: Missing user email")
                    continue
                if not program:
                    results['errors'].append(f"Row {row_num}: Missing program")
                    continue
                if not role:
                    results['errors'].append(f"Row {row_num}: Missing role")
                    continue
                if not granted_by:
                    results['errors'].append(f"Row {row_num}: Missing granted by")
                    continue

                # Validate role
                if role not in valid_roles:
                    results['errors'].append(f"Row {row_num}: Invalid role: {role}")
                    continue

                # Look up user
                user = self.am.get_user(email=user_email)
                if not user:
                    results['errors'].append(f"Row {row_num}: User not found: {user_email}")
                    continue

                # Look up program
                try:
                    program_id = self.am._resolve_program_id(program)
                except ValueError:
                    results['errors'].append(f"Row {row_num}: Program not found: {program}")
                    continue

                # Extract optional fields
                clinic = self._get_cell_value(ws, row_num, column_map.get('clinic'))
                location = self._get_cell_value(ws, row_num, column_map.get('location'))
                granted_date = self._get_cell_value(ws, row_num, column_map.get('granted_date'))
                reason = self._get_cell_value(ws, row_num, column_map.get('reason'))
                ticket = self._get_cell_value(ws, row_num, column_map.get('ticket'))
                review_cycle = self._get_cell_value(ws, row_num, column_map.get('review_cycle')) or 'Quarterly'

                # Validate review cycle
                if review_cycle not in ('Quarterly', 'Annual'):
                    review_cycle = 'Quarterly'

                # Resolve clinic_id if provided
                clinic_id = None
                if clinic:
                    try:
                        clinic_id = self.am._resolve_clinic_id(clinic, program_id)
                    except ValueError:
                        results['errors'].append(f"Row {row_num}: Clinic not found: {clinic}")
                        continue

                # Resolve location_id if provided
                location_id = None
                if location:
                    if not clinic_id:
                        results['errors'].append(f"Row {row_num}: Location requires clinic")
                        continue
                    try:
                        location_id = self.am._resolve_location_id(location, clinic_id)
                    except ValueError:
                        results['errors'].append(f"Row {row_num}: Location not found: {location}")
                        continue

                # Check for segregation of duties conflicts
                conflict_check = self.am.check_segregation_of_duties(
                    user['user_id'], role, program_id
                )
                if conflict_check['has_conflict']:
                    for conflict in conflict_check['conflicts']:
                        if conflict['severity'] == 'Block':
                            results['errors'].append(
                                f"Row {row_num}: Blocked - {conflict['existing_role']}+{role} conflict"
                            )
                            continue
                        else:
                            results['conflicts'].append(
                                f"Row {row_num}: Warning - {conflict['existing_role']}+{role} for {user_email}"
                            )

                # Skip if blocked
                if any(f"Row {row_num}: Blocked" in e for e in results['errors']):
                    continue

                # Grant access (unless dry run)
                if not dry_run:
                    self.am.grant_access(
                        user_id=user['user_id'],
                        program_id=program_id,
                        role=role,
                        granted_by=f"import:{granted_by}",
                        clinic_id=clinic_id,
                        location_id=location_id,
                        reason=reason or "Imported from Excel",
                        ticket=ticket,
                        review_cycle=review_cycle
                    )

                results['imported'] += 1

            except Exception as e:
                results['errors'].append(f"Row {row_num}: {str(e)}")

        return results

    def import_training(
        self,
        excel_path: str,
        sheet_name: str = None,
        dry_run: bool = False
    ) -> Dict[str, Any]:
        """
        Import training records from Excel.

        PURPOSE: Bulk import training completion records

        R EQUIVALENT:
            training <- read_excel("training.xlsx") %>%
              left_join(users, by = "email") %>%
              mutate(training_id = pmap_int(., record_training))

        PARAMETERS:
            excel_path: Path to Excel file
            sheet_name: Optional sheet name (defaults to first sheet)
            dry_run: If True, validate but don't actually import

        EXPECTED COLUMNS (flexible matching):
            - User Email (required): Email of existing user
            - Training Type (required): HIPAA Privacy, HIPAA Security, Part 11, SOC 2, PHI Handling,
                                       Cybersecurity, Application Training
            - Completed Date (optional): When training was completed
            - Expires Date: When training expires (defaults to 1 year from completion)
            - Certificate: Certificate reference number
            - Status: Current/Expired/Pending (calculated from dates if not provided)

        RETURNS:
            Dict with:
            - imported: Count of successfully imported records
            - skipped: Count of skipped rows
            - errors: List of error messages with row numbers

        AVIATION ANALOGY:
            Like importing ground school completion records. Each row shows
            which crew member completed which course and when their certificate
            expires.
        """
        wb = load_workbook(excel_path, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active

        # Find column mappings
        column_map = self._find_columns(ws, self.TRAINING_COLUMNS)

        # Validate required columns
        # Only user_email and training_type are required - completed_date is optional
        # because we track training for internal team (Propel Health) but not for
        # client employees whose organizations handle their own training compliance
        required = ['user_email', 'training_type']
        missing = [col for col in required if col not in column_map]
        if missing:
            return {
                'imported': 0,
                'skipped': 0,
                'errors': [f"Missing required columns: {', '.join(missing)}"]
            }

        results = {
            'imported': 0,
            'skipped': 0,
            'errors': []
        }

        # Process each row (skip header)
        for row_num in range(2, ws.max_row + 1):
            try:
                # Extract values
                user_email = self._get_cell_value(ws, row_num, column_map.get('user_email'))
                training_type = self._get_cell_value(ws, row_num, column_map.get('training_type'))
                responsibility = self._get_cell_value(ws, row_num, column_map.get('responsibility'))
                completed_date = self._get_cell_value(ws, row_num, column_map.get('completed_date'))

                # Skip empty rows
                if not user_email and not training_type:
                    continue

                # Validate required fields (only user_email and training_type are required)
                if not user_email:
                    results['errors'].append(f"Row {row_num}: Missing user email")
                    continue
                if not training_type:
                    results['errors'].append(f"Row {row_num}: Missing training type")
                    continue

                # Validate training type against class constant
                if training_type not in self.VALID_TRAINING_TYPES:
                    results['errors'].append(f"Row {row_num}: Invalid training type: {training_type}")
                    continue

                # Validate responsibility if provided (optional field, defaults to 'Propel Health')
                if responsibility and responsibility not in self.VALID_TRAINING_RESPONSIBILITY:
                    results['errors'].append(f"Row {row_num}: Invalid responsibility: {responsibility}. Use 'Client' or 'Propel Health'")
                    continue
                if not responsibility:
                    responsibility = 'Propel Health'  # Default to internal tracking

                # Look up user
                user = self.am.get_user(email=user_email)
                if not user:
                    results['errors'].append(f"Row {row_num}: User not found: {user_email}")
                    continue

                # Parse completed date if provided (optional)
                completed_date_str = None
                if completed_date:
                    completed_date_str = self._parse_date(completed_date)
                    if not completed_date_str:
                        results['errors'].append(f"Row {row_num}: Invalid date format: {completed_date}")
                        continue

                # Extract optional fields
                expires_date = self._get_cell_value(ws, row_num, column_map.get('expires_date'))
                certificate = self._get_cell_value(ws, row_num, column_map.get('certificate'))

                # Parse expires date if provided
                expires_date_str = None
                if expires_date:
                    expires_date_str = self._parse_date(expires_date)

                # Import training (unless dry run)
                if not dry_run:
                    # First assign the training with responsibility tracking
                    training_id = self.am.assign_training(
                        user_id=user['user_id'],
                        training_type=training_type,
                        assigned_by='import',
                        responsibility=responsibility
                    )

                    # Only complete training if completed_date was provided
                    # For client employees, we just track assignment - their org handles completion
                    if completed_date_str:
                        # Calculate expires_in_days if expires_date provided
                        expires_in_days = 365
                        if expires_date_str:
                            # Calculate days between completed and expires
                            completed_dt = datetime.fromisoformat(completed_date_str).date()
                            expires_dt = datetime.fromisoformat(expires_date_str).date()
                            expires_in_days = (expires_dt - completed_dt).days
                            if expires_in_days < 0:
                                expires_in_days = 0

                        # Complete it with the imported dates
                        self.am.complete_training(
                            training_id=training_id,
                            completed_date=completed_date_str,
                            certificate_reference=certificate,
                            expires_in_days=expires_in_days
                        )

                results['imported'] += 1

            except Exception as e:
                results['errors'].append(f"Row {row_num}: {str(e)}")

        return results

    def import_from_template(
        self,
        excel_path: str,
        dry_run: bool = False
    ) -> Dict[str, Any]:
        """
        Import from multi-tab template with all data types.

        PURPOSE: Import users, access, and training from a single file

        PARAMETERS:
            excel_path: Path to Excel file with Users, Access, Training tabs
            dry_run: If True, validate but don't actually import

        EXPECTED TABS:
            - Users: User information
            - Access: Access grants
            - Training: Training records

        RETURNS:
            Dict with results for each tab

        PROCESS:
            Imports in order: Users first, then Access, then Training.
            This ensures users exist before granting them access or training.

        EXAMPLE:
            results = importer.import_from_template("clinic_data.xlsx")
            print(f"Users: {results['users']['imported']}")
            print(f"Access: {results['access']['imported']}")
            print(f"Training: {results['training']['imported']}")
        """
        wb = load_workbook(excel_path, data_only=True)
        sheet_names = wb.sheetnames

        results = {
            'users': {'imported': 0, 'skipped': 0, 'errors': []},
            'access': {'imported': 0, 'skipped': 0, 'errors': [], 'conflicts': []},
            'training': {'imported': 0, 'skipped': 0, 'errors': []}
        }

        # Import users first (if tab exists)
        users_sheet = self._find_sheet(sheet_names, ['Users', 'User', 'Staff', 'Employees'])
        if users_sheet:
            results['users'] = self.import_users(excel_path, sheet_name=users_sheet, dry_run=dry_run)

        # Import access second (if tab exists)
        access_sheet = self._find_sheet(sheet_names, ['Access', 'Access Grants', 'Permissions', 'Roles'])
        if access_sheet:
            results['access'] = self.import_access(excel_path, sheet_name=access_sheet, dry_run=dry_run)

        # Import training third (if tab exists)
        training_sheet = self._find_sheet(sheet_names, ['Training', 'Training Records', 'Certifications', 'Courses'])
        if training_sheet:
            results['training'] = self.import_training(excel_path, sheet_name=training_sheet, dry_run=dry_run)

        return results

    def generate_import_template(self, output_path: str) -> str:
        """
        Generate blank Excel template with correct columns and example data.

        PURPOSE: Create a template that users can fill in for import

        PARAMETERS:
            output_path: Where to save the template

        CREATES TABS:
            - Users: User information columns with example row
            - Access: Access grant columns with example row
            - Training: Training record columns with example row
            - Instructions: Column definitions and valid values

        FEATURES:
            - Data validation dropdowns for constrained fields (Role, Status, etc.)
            - Example rows showing expected data format
            - Color-coded required vs optional columns
            - Instructions sheet with field descriptions

        RETURNS:
            str: Path to generated template

        EXAMPLE:
            path = importer.generate_import_template("Template.xlsx")
            print(f"Template saved to: {path}")
        """
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Create Users sheet
        self._create_users_sheet(wb)

        # Create Access sheet
        self._create_access_sheet(wb)

        # Create Training sheet
        self._create_training_sheet(wb)

        # Create Instructions sheet
        self._create_instructions_sheet(wb)

        # Save
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))

        return str(output_path)

    # =========================================================================
    # TEMPLATE SHEET CREATORS
    # =========================================================================

    def _create_users_sheet(self, wb: Workbook) -> None:
        """Create the Users sheet with headers, example, and validation."""
        ws = wb.create_sheet("Users")

        # Headers
        headers = [
            ('Name', True, 25),           # (name, required, width)
            ('Email', True, 30),
            ('Organization', False, 20),
            ('Is Business Associate', False, 20),
            ('Status', False, 15),
            ('Notes', False, 40)
        ]

        for col, (header, required, width) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            if required:
                cell.font = Font(color="FFFFFF", bold=True, size=11, underline='single')
            ws.column_dimensions[get_column_letter(col)].width = width

        # Example row
        example_data = [
            'John Smith',
            'jsmith@clinic.com',
            'Internal',
            'No',
            'Active',
            'New hire - started Dec 2025'
        ]
        for col, value in enumerate(example_data, 1):
            cell = ws.cell(row=2, column=col, value=value)
            cell.fill = self.EXAMPLE_FILL

        # Data validation for Is Business Associate
        ba_validation = DataValidation(
            type="list",
            formula1='"Yes,No"',
            allow_blank=True
        )
        ws.add_data_validation(ba_validation)
        ba_validation.add(f"D3:D1000")

        # Data validation for Status
        status_validation = DataValidation(
            type="list",
            formula1='"Active,Inactive,Terminated"',
            allow_blank=True
        )
        ws.add_data_validation(status_validation)
        status_validation.add(f"E3:E1000")

        ws.freeze_panes = 'A2'

    def _create_access_sheet(self, wb: Workbook) -> None:
        """Create the Access sheet with headers, example, and validation."""
        ws = wb.create_sheet("Access")

        # Headers
        headers = [
            ('User Email', True, 30),
            ('Program', True, 20),
            ('Clinic', False, 25),
            ('Location', False, 25),
            ('Role', True, 15),
            ('Granted Date', False, 15),
            ('Granted By', True, 20),
            ('Reason', False, 40),
            ('Ticket', False, 15),
            ('Review Cycle', False, 15)
        ]

        for col, (header, required, width) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            if required:
                cell.font = Font(color="FFFFFF", bold=True, size=11, underline='single')
            ws.column_dimensions[get_column_letter(col)].width = width

        # Example row
        example_data = [
            'jsmith@clinic.com',
            'P4M',
            'Portland',
            'Breast Surgery West',
            'Coordinator',
            '2025-12-19',
            'Manager Name',
            'New hire needs access for scheduling',
            'SR-12345',
            'Quarterly'
        ]
        for col, value in enumerate(example_data, 1):
            cell = ws.cell(row=2, column=col, value=value)
            cell.fill = self.EXAMPLE_FILL

        # Data validation for Role
        role_validation = DataValidation(
            type="list",
            formula1='"Read-Only,Read-Write,Read-Write-Order,Clinic-Manager,Analytics-Only,Admin,Auditor"',
            allow_blank=False
        )
        ws.add_data_validation(role_validation)
        role_validation.add(f"E3:E1000")

        # Data validation for Review Cycle
        review_validation = DataValidation(
            type="list",
            formula1='"Quarterly,Annual"',
            allow_blank=True
        )
        ws.add_data_validation(review_validation)
        review_validation.add(f"J3:J1000")

        ws.freeze_panes = 'A2'

    def _create_training_sheet(self, wb: Workbook) -> None:
        """Create the Training sheet with headers, example, and validation."""
        ws = wb.create_sheet("Training")

        # Headers
        # Only User Email and Training Type are required
        # Responsibility, Completed Date, Expires Date, Certificate are all optional
        headers = [
            ('User Email', True, 30),
            ('Training Type', True, 20),
            ('Training Responsibility', False, 22),  # Client or Propel Health
            ('Completed Date', False, 15),   # Optional - only track for internal team
            ('Expires Date', False, 15),
            ('Certificate', False, 20)
        ]

        for col, (header, required, width) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            if required:
                cell.font = Font(color="FFFFFF", bold=True, size=11, underline='single')
            ws.column_dimensions[get_column_letter(col)].width = width

        # Example row
        example_data = [
            'jsmith@clinic.com',
            'HIPAA',
            'Propel Health',
            '2025-12-15',
            '2026-12-15',
            'CERT-12345'
        ]
        for col, value in enumerate(example_data, 1):
            cell = ws.cell(row=2, column=col, value=value)
            cell.fill = self.EXAMPLE_FILL

        # Data validation for Training Type
        # Active: HIPAA, Cybersecurity, Application Training
        # Reserved: SOC 2, HITRUST, Part 11
        type_validation = DataValidation(
            type="list",
            formula1='"HIPAA,Cybersecurity,Application Training,SOC 2,HITRUST,Part 11"',
            allow_blank=False
        )
        ws.add_data_validation(type_validation)
        type_validation.add(f"B3:B1000")

        # Data validation for Training Responsibility
        responsibility_validation = DataValidation(
            type="list",
            formula1='"Propel Health,Client"',
            allow_blank=True
        )
        ws.add_data_validation(responsibility_validation)
        responsibility_validation.add(f"C3:C1000")

        ws.freeze_panes = 'A2'

    def _create_instructions_sheet(self, wb: Workbook) -> None:
        """Create the Instructions sheet with field definitions."""
        ws = wb.create_sheet("Instructions")

        # Title
        ws['A1'] = "Import Template Instructions"
        ws['A1'].font = Font(size=16, bold=True)

        # General instructions
        ws['A3'] = "How to Use This Template:"
        ws['A3'].font = Font(bold=True)
        instructions = [
            "1. Fill in the Users tab first (users must exist before granting access or training)",
            "2. Fill in the Access tab to grant access to users",
            "3. Fill in the Training tab to record training completions",
            "4. Delete the green example rows before importing",
            "5. Save the file and run: python3 run.py --import-access-template 'filename.xlsx'",
            "",
            "Required columns are underlined in the headers.",
            "Green rows are examples - delete them before importing.",
            "Dropdown lists are provided for constrained fields."
        ]
        for i, line in enumerate(instructions, 4):
            ws[f'A{i}'] = line

        # Field definitions
        row = 15
        ws[f'A{row}'] = "Field Definitions:"
        ws[f'A{row}'].font = Font(bold=True)

        row += 2
        definitions = [
            ("Users Tab:", ""),
            ("  Name", "Full name (e.g., 'John Smith')"),
            ("  Email", "Email address - must be unique"),
            ("  Organization", "'Internal' for employees, company name for contractors"),
            ("  Is Business Associate", "Yes if external party requiring HIPAA BAA"),
            ("  Status", "Active, Inactive, or Terminated"),
            ("", ""),
            ("Access Tab:", ""),
            ("  User Email", "Email of an existing user"),
            ("  Program", "Program prefix (e.g., 'P4M') or full name"),
            ("  Clinic", "Optional - limit access to specific clinic"),
            ("  Location", "Optional - limit access to specific location (requires clinic)"),
            ("  Role", "Read-Only, Read-Write, Read-Write-Order, Clinic-Manager, Analytics-Only, Admin, or Auditor"),
            ("  Granted Date", "Date access was granted (YYYY-MM-DD)"),
            ("  Granted By", "Who approved the access"),
            ("  Reason", "Business justification for access"),
            ("  Ticket", "Reference to approval ticket/email"),
            ("  Review Cycle", "Quarterly (90 days) or Annual (365 days)"),
            ("", ""),
            ("Training Tab:", ""),
            ("  User Email", "Email of an existing user (required)"),
            ("  Training Type", "Active: HIPAA, Cybersecurity, Application Training. Reserved: SOC 2, HITRUST, Part 11 (required)"),
            ("  Training Responsibility", "'Client' if client org tracks training, 'Propel Health' if PHP tracks (defaults to Propel Health)"),
            ("  Completed Date", "Optional - date training was completed (YYYY-MM-DD). Only for internal team."),
            ("  Expires Date", "Optional - when training expires (defaults to 1 year from completion)"),
            ("  Certificate", "Optional - certificate or reference number"),
        ]

        for field, desc in definitions:
            ws[f'A{row}'] = field
            ws[f'B{row}'] = desc
            row += 1

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 60

    # =========================================================================
    # HELPER METHODS
    # =========================================================================

    def _find_columns(
        self,
        ws,
        column_mappings: Dict[str, List[str]]
    ) -> Dict[str, int]:
        """
        Find column indices by matching header names.

        PURPOSE: Map flexible column names to column indices

        PARAMETERS:
            ws: Worksheet to search
            column_mappings: Dict of field_name -> [possible_column_names]

        RETURNS:
            Dict of field_name -> column_index (1-based)

        WHY THIS APPROACH:
            Users create Excel files with different column names. By checking
            multiple possible names, we're forgiving of variations like
            "Email" vs "Email Address" vs "E-mail".
        """
        found_columns = {}

        # Read header row
        headers = {}
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value:
                headers[col] = str(cell_value).strip().lower()

        # Match each field to a column
        for field_name, possible_names in column_mappings.items():
            for possible_name in possible_names:
                possible_lower = possible_name.lower()
                for col, header in headers.items():
                    if header == possible_lower:
                        found_columns[field_name] = col
                        break
                if field_name in found_columns:
                    break

        return found_columns

    def _get_cell_value(
        self,
        ws,
        row: int,
        col: int
    ) -> Optional[str]:
        """
        Get cell value as string, handling None and various types.

        PURPOSE: Safely extract cell values with type conversion

        PARAMETERS:
            ws: Worksheet
            row: Row number (1-based)
            col: Column number (1-based)

        RETURNS:
            String value or None if empty
        """
        if col is None:
            return None

        value = ws.cell(row=row, column=col).value

        if value is None:
            return None

        # Handle datetime objects
        if isinstance(value, datetime):
            return value.strftime('%Y-%m-%d')

        # Handle date objects
        if isinstance(value, date):
            return value.strftime('%Y-%m-%d')

        # Convert to string and strip whitespace
        return str(value).strip()

    def _parse_boolean(self, value: Optional[str]) -> bool:
        """
        Parse a boolean value from various string representations.

        PURPOSE: Handle Yes/No, True/False, 1/0, etc.

        PARAMETERS:
            value: String to parse

        RETURNS:
            bool: True if value represents true, False otherwise
        """
        if value is None:
            return False

        value_lower = str(value).lower().strip()
        return value_lower in ('yes', 'true', '1', 'y', 'x')

    def _parse_date(self, value: Any) -> Optional[str]:
        """
        Parse a date value to ISO format string.

        PURPOSE: Handle various date formats from Excel

        PARAMETERS:
            value: Date value (could be datetime, date, or string)

        RETURNS:
            str: ISO format date (YYYY-MM-DD) or None if invalid
        """
        if value is None:
            return None

        # Already a datetime or date
        if isinstance(value, datetime):
            return value.strftime('%Y-%m-%d')
        if isinstance(value, date):
            return value.strftime('%Y-%m-%d')

        # Try to parse string
        value_str = str(value).strip()

        # Common formats to try
        formats = [
            '%Y-%m-%d',      # 2025-12-19
            '%m/%d/%Y',      # 12/19/2025
            '%d/%m/%Y',      # 19/12/2025
            '%Y/%m/%d',      # 2025/12/19
            '%m-%d-%Y',      # 12-19-2025
            '%d-%m-%Y',      # 19-12-2025
            '%B %d, %Y',     # December 19, 2025
            '%b %d, %Y',     # Dec 19, 2025
        ]

        for fmt in formats:
            try:
                dt = datetime.strptime(value_str, fmt)
                return dt.strftime('%Y-%m-%d')
            except ValueError:
                continue

        return None

    def _is_valid_email(self, email: str) -> bool:
        """
        Validate email format.

        PURPOSE: Basic email validation

        PARAMETERS:
            email: Email address to validate

        RETURNS:
            bool: True if email appears valid
        """
        # Simple regex for email validation
        pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        return bool(re.match(pattern, email))

    def _find_sheet(
        self,
        sheet_names: List[str],
        possible_names: List[str]
    ) -> Optional[str]:
        """
        Find a sheet by checking multiple possible names.

        PURPOSE: Flexible sheet name matching

        PARAMETERS:
            sheet_names: List of sheet names in workbook
            possible_names: List of possible sheet names to look for

        RETURNS:
            str: Matching sheet name, or None if not found
        """
        for possible in possible_names:
            for sheet in sheet_names:
                if sheet.lower() == possible.lower():
                    return sheet
        return None


# =============================================================================
# MODULE TEST
# =============================================================================

if __name__ == "__main__":
    print("Testing AccessImporter...")

    if not OPENPYXL_AVAILABLE:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    else:
        am = AccessManager()
        importer = AccessImporter(am)

        # Generate template
        template_path = "outputs/test_template.xlsx"
        importer.generate_import_template(template_path)
        print(f"Generated template: {template_path}")
