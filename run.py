#!/usr/bin/env python3
"""
CONFIGURATIONS TOOLKIT CLI

PURPOSE: Command-line interface for managing system configurations
         with inheritance across Program → Clinic → Location hierarchy

R EQUIVALENT: Like an R package's main script that routes to different
              functions based on command-line arguments

AVIATION ANALOGY: Like a flight management system interface -
                  different modes for different operations

USAGE EXAMPLES:
    # Initialize database and load definitions
    python3 run.py --init

    # Import from Word document
    python3 run.py --import "Portland_Clinic_Spec.docx" --program P4M

    # View configs for a clinic
    python3 run.py --view --program P4M --clinic Portland

    # View effective config for a location
    python3 run.py --view --program P4M --clinic Portland --location "Breast Surgery West"

    # Set a config value
    python3 run.py --set helpdesk_phone --value "503.216.6407" --program P4M --clinic Portland --location "Breast Surgery West"

    # Update provider NPI
    python3 run.py --update-provider "Christine Kemp" --npi "1215158639"

    # View audit trail for a config
    python3 run.py --audit helpdesk_phone --program P4M --clinic Portland

    # Export all configs to Excel
    python3 run.py --export --program P4M --output excel

    # Compare clinic to defaults
    python3 run.py --compare --program P4M --clinic Portland

    # Create new program
    python3 run.py --create-program "Discover" --type attached --attach-to "P4M,PRE,GRX"

    # List all programs and their hierarchy
    python3 run.py --list-programs

AUTHOR: Glen Lewis
DATE: 2024
"""

import json
import sys
import os
from pathlib import Path
from datetime import datetime

# Add project root to path for imports
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# Import CLI components from cli/ package
from cli import create_parser

# Import managers and formatters
from database.config_manager import ConfigurationManager
from managers.inheritance_manager import InheritanceManager
from managers.update_manager import QuickUpdateManager
from managers.access_manager import AccessManager
from managers.access_import import AccessImporter
from formatters.config_excel_formatter import ConfigExcelFormatter
from formatters.access_excel_formatter import AccessExcelFormatter
from reports.compliance_reports import ComplianceReports


# ============================================================================
# COMMAND HANDLERS
# ============================================================================

def handle_init(cm: ConfigurationManager, args) -> None:
    """Initialize database and load definitions."""
    print("Initializing Configurations Toolkit...")
    cm.initialize_schema()
    count = cm.load_definitions_from_yaml()
    print(f"Loaded {count} configuration definitions")
    print("Initialization complete!")


def handle_import(cm: ConfigurationManager, args) -> None:
    """
    Import configurations from Word document.

    When --reimport is specified, clears existing program data first to allow
    a fresh import without duplicate entries. Useful after parser fixes.
    """
    from parsers.word_parser import ClinicSpecParser

    if not args.program:
        print("Error: --program required for import")
        sys.exit(1)

    program = cm.get_program_by_prefix(args.program)
    if not program:
        print(f"Error: Program not found: {args.program}")
        sys.exit(1)

    # Handle --reimport: clear existing data before importing
    if args.reimport:
        print(f"\nClearing existing data for {args.program}...")
        clear_counts = cm.clear_program_data(program['program_id'], keep_structure=False)
        print(f"  Cleared: {clear_counts['config_values']} configs, "
              f"{clear_counts['locations']} locations, "
              f"{clear_counts['providers']} providers")

    print(f"\nParsing: {args.import_file}")
    parser = ClinicSpecParser(args.import_file)
    result = parser.parse()

    print(f"\nParsed document:")
    print(f"  Clinic: {result.get('clinic_name')}")
    print(f"  Locations: {len(result.get('scope_locations', []))}")
    print(f"  Configurations: {len(result.get('configurations', []))}")
    print(f"  Mapped configs: {len(result.get('mapped_configs', []))}")
    print(f"  Providers: {len(result.get('providers', []))}")

    print("\nImporting to database...")
    counts = cm.import_from_parsed_doc(result, program['program_id'],
                                       source_document=args.import_file)

    print(f"\nImport complete:")
    print(f"  Clinics created: {counts['clinics']}")
    print(f"  Locations created: {counts['locations']}")
    print(f"  Location configs: {counts['location_configs']}")
    print(f"  Providers added: {counts['providers']}")


def handle_view(cm: ConfigurationManager, args) -> None:
    """View configurations."""
    if not args.program:
        print("Error: --program required for view")
        sys.exit(1)

    program = cm.get_program_by_prefix(args.program)
    if not program:
        print(f"Error: Program not found: {args.program}")
        sys.exit(1)

    program_id = program['program_id']
    clinic_id = None
    location_id = None

    # Get clinic if specified
    if args.clinic:
        clinic = cm.get_clinic_by_name(program_id, args.clinic)
        if not clinic:
            print(f"Error: Clinic not found: {args.clinic}")
            sys.exit(1)
        clinic_id = clinic['clinic_id']

    # Get location if specified
    if args.location and clinic_id:
        location = cm.get_location_by_name(clinic_id, args.location)
        if not location:
            print(f"Error: Location not found: {args.location}")
            sys.exit(1)
        location_id = location['location_id']

    # Get effective configuration
    configs = cm.get_effective_config(program_id, clinic_id, location_id)

    # Group by category
    by_category = {}
    cursor = cm.conn.cursor()
    cursor.execute("SELECT * FROM config_definitions ORDER BY category, display_order")

    for defn in cursor.fetchall():
        key = defn['config_key']
        category = defn['category']

        if category not in by_category:
            by_category[category] = []

        config = configs.get(key, {})
        by_category[category].append({
            'key': key,
            'name': defn['display_name'],
            'value': config.get('value'),
            'level': config.get('effective_level'),
            'is_override': config.get('is_override', False)
        })

    # Print results
    context = [program['name']]
    if args.clinic:
        context.append(args.clinic)
    if args.location:
        context.append(args.location)

    print(f"\nConfigurations for: {' > '.join(context)}")
    print("=" * 60)

    for category, items in by_category.items():
        print(f"\n{category.upper().replace('_', ' ')}")
        print("-" * 40)

        for item in items:
            override_marker = "*" if item['is_override'] else " "
            value = item['value'] or "(not set)"
            level = f"[{item['level']}]" if item['level'] else ""

            # Truncate long values
            if len(str(value)) > 40:
                value = str(value)[:37] + "..."

            print(f"  {override_marker} {item['name']}: {value} {level}")

    print("\n* = Override from parent level")


def handle_list_programs(cm: ConfigurationManager, args) -> None:
    """List all programs with hierarchy."""
    cursor = cm.conn.cursor()
    cursor.execute("""
        SELECT * FROM programs ORDER BY name
    """)

    programs = cursor.fetchall()

    print("\nPrograms:")
    print("=" * 60)

    for prog in programs:
        print(f"\n{prog['name']} ({prog['prefix']})")
        print(f"  Type: {prog.get('program_type', 'clinic_based')}")
        print(f"  ID: {prog['program_id']}")

        # Get clinics
        cursor.execute("""
            SELECT * FROM clinics WHERE program_id = ? ORDER BY name
        """, (prog['program_id'],))

        clinics = cursor.fetchall()
        if clinics:
            print("  Clinics:")
            for clinic in clinics:
                print(f"    └── {clinic['name']} ({clinic.get('code', 'N/A')})")

                # Get locations
                cursor.execute("""
                    SELECT * FROM locations WHERE clinic_id = ? ORDER BY name
                """, (clinic['clinic_id'],))

                for loc in cursor.fetchall():
                    print(f"        └── {loc['name']}")


def handle_audit(cm: ConfigurationManager, args) -> None:
    """View audit history for a config key."""
    if not args.program:
        print("Error: --program required for audit")
        sys.exit(1)

    program = cm.get_program_by_prefix(args.program)
    if not program:
        print(f"Error: Program not found: {args.program}")
        sys.exit(1)

    program_id = program['program_id']
    clinic_id = None
    location_id = None

    if args.clinic:
        clinic = cm.get_clinic_by_name(program_id, args.clinic)
        if clinic:
            clinic_id = clinic['clinic_id']

    if args.location and clinic_id:
        location = cm.get_location_by_name(clinic_id, args.location)
        if location:
            location_id = location['location_id']

    history = cm.get_config_history(args.audit, program_id, clinic_id, location_id)

    print(f"\nAudit History for: {args.audit}")
    print("=" * 60)

    if not history:
        print("No history found")
        return

    for entry in history:
        print(f"\n{entry['changed_date']}")
        print(f"  Old: {entry['old_value']}")
        print(f"  New: {entry['new_value']}")
        print(f"  By: {entry['changed_by']}")
        if entry['change_reason']:
            print(f"  Reason: {entry['change_reason']}")


def handle_set(cm: ConfigurationManager, args) -> None:
    """
    Set a configuration value.

    Can set at program level or location level.
    Location can be specified directly without requiring --clinic.
    """
    if not args.program:
        print("Error: --program required")
        sys.exit(1)

    if not args.value:
        print("Error: --value required")
        sys.exit(1)

    program = cm.get_program_by_prefix(args.program)
    if not program:
        print(f"Error: Program not found: {args.program}")
        sys.exit(1)

    program_id = program['program_id']
    clinic_id = None
    location_id = None

    # If --location specified, find it (may need to search across all clinics)
    if args.location:
        # Get all clinics for this program
        cursor = cm.conn.cursor()
        cursor.execute("SELECT clinic_id FROM clinics WHERE program_id = ?", (program_id,))
        clinics = cursor.fetchall()

        # Search for the location across all clinics
        for clinic_row in clinics:
            c_id = clinic_row['clinic_id']
            location = cm.get_location_by_name(c_id, args.location)
            if location:
                clinic_id = c_id
                location_id = location['location_id']
                break

        if not location_id:
            print(f"Error: Location not found: {args.location}")
            sys.exit(1)

    elif args.clinic:
        # Just clinic specified (legacy support)
        clinic = cm.get_clinic_by_name(program_id, args.clinic)
        if clinic:
            clinic_id = clinic['clinic_id']
        else:
            print(f"Error: Clinic not found: {args.clinic}")
            sys.exit(1)

    cm.set_config(
        args.set,
        args.value,
        program_id,
        clinic_id,
        location_id,
        source='cli',
        rationale='Set via CLI'
    )

    level = "location" if location_id else ("clinic" if clinic_id else "program")
    print(f"Set {args.set} = {args.value} at {level} level")


def handle_update_provider(cm: ConfigurationManager, args) -> None:
    """Update provider NPI."""
    if not args.npi:
        print("Error: --npi required with --update-provider")
        sys.exit(1)

    qm = QuickUpdateManager(cm)
    count = qm.update_provider_npi(
        args.update_provider,
        args.npi,
        program_prefix=args.program
    )

    print(f"Updated {count} provider(s)")


def handle_export(cm: ConfigurationManager, args) -> None:
    """Export configurations to Excel."""
    if not args.program:
        print("Error: --program required for export")
        sys.exit(1)

    program = cm.get_program_by_prefix(args.program)
    if not program:
        print(f"Error: Program not found: {args.program}")
        sys.exit(1)

    # Determine output path
    if args.output:
        output_path = args.output
    else:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        output_path = f"outputs/{args.program}_configs_{timestamp}.xlsx"

    formatter = ConfigExcelFormatter(cm)

    if args.clinic:
        output_path = formatter.export_clinic(args.program, args.clinic, output_path)
    else:
        output_path = formatter.export_program(args.program, output_path)

    print(f"Exported to: {output_path}")


def handle_compare(cm: ConfigurationManager, args) -> None:
    """Compare clinic/location to defaults."""
    if not args.program:
        print("Error: --program required")
        sys.exit(1)

    program = cm.get_program_by_prefix(args.program)
    if not program:
        print(f"Error: Program not found: {args.program}")
        sys.exit(1)

    program_id = program['program_id']
    clinic_id = None
    location_id = None

    if args.clinic:
        clinic = cm.get_clinic_by_name(program_id, args.clinic)
        if clinic:
            clinic_id = clinic['clinic_id']

    if args.location and clinic_id:
        location = cm.get_location_by_name(clinic_id, args.location)
        if location:
            location_id = location['location_id']

    differences = cm.compare_to_defaults(program_id, clinic_id, location_id)

    print(f"\nDifferences from defaults:")
    print("=" * 60)

    if not differences:
        print("No differences found - all values match defaults")
        return

    for diff in differences:
        print(f"\n{diff['config_key']}")
        print(f"  Default: {diff['default_value']}")
        print(f"  Current: {diff['current_value']} [{diff['level']}]")


def handle_tree(cm: ConfigurationManager, args) -> None:
    """Show inheritance tree for a config."""
    if not args.program:
        print("Error: --program required")
        sys.exit(1)

    program = cm.get_program_by_prefix(args.program)
    if not program:
        print(f"Error: Program not found: {args.program}")
        sys.exit(1)

    im = InheritanceManager(cm)
    tree_str = im.print_inheritance_tree(args.tree, program['program_id'])
    print(tree_str)


def handle_create_program(cm: ConfigurationManager, args) -> None:
    """Create a new program."""
    if not args.prefix:
        print("Error: --prefix required for --create-program")
        sys.exit(1)

    program_id = cm.create_program(
        args.create_program,
        args.prefix,
        program_type=args.type
    )

    # Handle attach-to
    if args.attach_to:
        parent_prefixes = [p.strip() for p in args.attach_to.split(',')]
        for parent_prefix in parent_prefixes:
            parent = cm.get_program_by_prefix(parent_prefix)
            if parent:
                cm.attach_program(parent['program_id'], program_id)
                print(f"Attached to: {parent['name']}")
            else:
                print(f"Warning: Parent program not found: {parent_prefix}")

    print(f"\nCreated program: {args.create_program} ({args.prefix})")


def handle_create_clinic(cm: ConfigurationManager, args) -> None:
    """Create a new clinic."""
    if not args.program:
        print("Error: --program required")
        sys.exit(1)

    program = cm.get_program_by_prefix(args.program)
    if not program:
        print(f"Error: Program not found: {args.program}")
        sys.exit(1)

    clinic_id = cm.create_clinic(
        program['program_id'],
        args.create_clinic,
        code=args.code
    )

    print(f"Created clinic: {args.create_clinic} ({clinic_id})")


def handle_create_location(cm: ConfigurationManager, args) -> None:
    """Create a new location."""
    if not args.program or not args.clinic:
        print("Error: --program and --clinic required")
        sys.exit(1)

    program = cm.get_program_by_prefix(args.program)
    if not program:
        print(f"Error: Program not found: {args.program}")
        sys.exit(1)

    clinic = cm.get_clinic_by_name(program['program_id'], args.clinic)
    if not clinic:
        print(f"Error: Clinic not found: {args.clinic}")
        sys.exit(1)

    location_id = cm.create_location(
        clinic['clinic_id'],
        args.create_location,
        code=args.code
    )

    print(f"Created location: {args.create_location} ({location_id})")


# ============================================================================
# ACCESS MANAGEMENT HANDLERS
# ============================================================================
# Handlers for Part 11, HIPAA, and SOC 2 compliance tracking

def handle_init_access(am: AccessManager, args) -> None:
    """Initialize access management schema."""
    print("Initializing access management schema...")
    am.initialize_schema()
    print("Access schema initialized!")


def handle_add_user(am: AccessManager, args) -> None:
    """Create a new user."""
    if not args.email:
        print("Error: --email required with --add-user")
        sys.exit(1)

    try:
        user_id = am.create_user(
            name=args.add_user,
            email=args.email,
            organization=args.organization,
            is_business_associate=args.business_associate
        )
        print(f"\nCreated user: {args.add_user}")
        print(f"  User ID: {user_id}")
        print(f"  Email: {args.email}")
        print(f"  Organization: {args.organization}")
        if args.business_associate:
            print("  Business Associate: Yes (HIPAA BAA required)")
    except ValueError as e:
        print(f"Error: {e}")
        sys.exit(1)


def handle_list_users(am: AccessManager, args) -> None:
    """List all users."""
    users = am.list_users(
        status_filter=args.status,
        organization_filter=args.organization if args.organization != 'Internal' else None
    )

    if not users:
        print("No users found")
        return

    print(f"\nUsers ({len(users)} total)")
    print("=" * 80)

    for user in users:
        status_marker = ""
        if user['status'] == 'Terminated':
            status_marker = " [TERMINATED]"
        elif user['status'] == 'Inactive':
            status_marker = " [INACTIVE]"

        ba_marker = " (BA)" if user['is_business_associate'] else ""

        print(f"\n{user['name']}{status_marker}{ba_marker}")
        print(f"  ID: {user['user_id']}")
        print(f"  Email: {user['email']}")
        print(f"  Organization: {user['organization']}")
        if 'active_access_count' in user:
            print(f"  Active Access: {user['active_access_count']} grants")


def handle_terminate_user(am: AccessManager, args) -> None:
    """Terminate a user and revoke all access."""
    if not args.reason:
        print("Error: --reason required with --terminate-user")
        sys.exit(1)

    if not args.by:
        print("Error: --by required with --terminate-user")
        sys.exit(1)

    try:
        result = am.terminate_user(
            user_id=args.terminate_user,
            reason=args.reason,
            terminated_by=args.by
        )

        print(f"\nUser terminated: {result['user_id']}")
        print(f"  Terminated by: {args.by}")
        print(f"  Reason: {args.reason}")
        print(f"  Access grants revoked: {result['access_revoked']}")
    except ValueError as e:
        print(f"Error: {e}")
        sys.exit(1)


def handle_grant_access(am: AccessManager, cm: ConfigurationManager, args) -> None:
    """Grant access to a user."""
    if not args.user:
        print("Error: --user required with --grant-access")
        sys.exit(1)

    if not args.program:
        print("Error: --program required with --grant-access")
        sys.exit(1)

    if not args.role:
        print("Error: --role required with --grant-access")
        sys.exit(1)

    if not args.by:
        print("Error: --by required with --grant-access")
        sys.exit(1)

    try:
        access_id = am.grant_access(
            user_id=args.user,
            program_id=args.program,
            role=args.role,
            granted_by=args.by,
            clinic_id=args.clinic,
            location_id=args.location,
            reason=args.reason
        )

        scope_parts = [args.program]
        if args.clinic:
            scope_parts.append(args.clinic)
        if args.location:
            scope_parts.append(args.location)

        print(f"\nAccess granted!")
        print(f"  Access ID: {access_id}")
        print(f"  User: {args.user}")
        print(f"  Role: {args.role}")
        print(f"  Scope: {' > '.join(scope_parts)}")
        print(f"  Granted by: {args.by}")
    except ValueError as e:
        print(f"Error: {e}")
        sys.exit(1)


def handle_revoke_access(am: AccessManager, args) -> None:
    """Revoke an access grant."""
    if not args.access_id:
        print("Error: --access-id required with --revoke-access")
        sys.exit(1)

    if not args.by:
        print("Error: --by required with --revoke-access")
        sys.exit(1)

    if not args.reason:
        print("Error: --reason required with --revoke-access")
        sys.exit(1)

    try:
        result = am.revoke_access(
            access_id=args.access_id,
            revoked_by=args.by,
            reason=args.reason
        )

        print(f"\nAccess revoked!")
        print(f"  Access ID: {args.access_id}")
        print(f"  User: {result['user_name']}")
        print(f"  Role: {result['role']}")
        print(f"  Revoked by: {args.by}")
        print(f"  Reason: {args.reason}")
    except ValueError as e:
        print(f"Error: {e}")
        sys.exit(1)


def handle_list_access(am: AccessManager, args) -> None:
    """List access grants."""
    if args.user:
        # List access for a specific user
        access_list = am.get_user_access(args.user, active_only=True)
        print(f"\nAccess for: {args.user}")
    elif args.program or args.clinic or args.location:
        # List access for a scope
        access_list = am.get_access_by_scope(
            program_id=args.program,
            clinic_id=args.clinic,
            location_id=args.location
        )
        scope_parts = []
        if args.program:
            scope_parts.append(args.program)
        if args.clinic:
            scope_parts.append(args.clinic)
        if args.location:
            scope_parts.append(args.location)
        print(f"\nAccess for scope: {' > '.join(scope_parts)}")
    else:
        print("Error: --user, --program, --clinic, or --location required")
        sys.exit(1)

    print("=" * 80)

    if not access_list:
        print("No access grants found")
        return

    for access in access_list:
        scope_parts = [access.get('program_name', access.get('program_prefix', ''))]
        if access.get('clinic_name'):
            scope_parts.append(access['clinic_name'])
        if access.get('location_name'):
            scope_parts.append(access['location_name'])

        user_name = access.get('user_name', '')
        print(f"\n[{access['access_id']}] {user_name} - {access['role']}")
        print(f"  Scope: {' > '.join(scope_parts)}")
        print(f"  Granted: {access['granted_date']} by {access.get('granted_by', 'unknown')}")
        if access.get('next_review_due'):
            print(f"  Next Review: {access['next_review_due']}")


def handle_reviews_due(am: AccessManager, args) -> None:
    """Show overdue access reviews."""
    reviews = am.get_reviews_due(program_id=args.program)

    print(f"\nOverdue Access Reviews")
    print("=" * 80)

    if not reviews:
        print("No reviews are overdue - all access reviews are current!")
        return

    print(f"Total overdue: {len(reviews)}")

    for review in reviews:
        days = int(review.get('days_overdue', 0))
        scope_parts = [review.get('program_name', '')]
        if review.get('clinic_name'):
            scope_parts.append(review['clinic_name'])
        if review.get('location_name'):
            scope_parts.append(review['location_name'])

        print(f"\n[{review['access_id']}] {review['user_name']} - {days} days overdue")
        print(f"  Email: {review['email']}")
        print(f"  Role: {review['role']}")
        print(f"  Scope: {' > '.join(scope_parts)}")
        print(f"  Due: {review['next_review_due']}")
        if review.get('last_review_date'):
            print(f"  Last Review: {review['last_review_date']}")


def handle_conduct_review(am: AccessManager, args) -> None:
    """Conduct an access review."""
    if not args.access_id:
        print("Error: --access-id required with --conduct-review")
        sys.exit(1)

    if not args.review_status:
        print("Error: --review-status required (Certified, Revoked, or Modified)")
        sys.exit(1)

    if not args.by:
        print("Error: --by required with --conduct-review")
        sys.exit(1)

    try:
        result = am.conduct_review(
            access_id=args.access_id,
            reviewed_by=args.by,
            status=args.review_status,
            notes=args.notes
        )

        print(f"\nReview recorded!")
        print(f"  Review ID: {result['review_id']}")
        print(f"  Access ID: {args.access_id}")
        print(f"  User: {result['user_name']}")
        print(f"  Decision: {result['status']}")
        print(f"  Reviewed by: {args.by}")
        if result.get('next_review_due'):
            print(f"  Next Review Due: {result['next_review_due']}")
    except ValueError as e:
        print(f"Error: {e}")
        sys.exit(1)


def handle_export_review_worksheet(am: AccessManager, args) -> None:
    """Export review worksheet to Excel."""
    # Get access needing review
    reviews = am.get_reviews_due(program_id=args.program)

    if not reviews:
        print("No reviews due - nothing to export")
        return

    # Determine output path
    if args.output:
        output_path = args.output
    else:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        output_path = f"outputs/review_worksheet_{timestamp}.xlsx"

    # Build scope description
    scope_parts = []
    if args.program:
        scope_parts.append(f"Program: {args.program}")
    if args.clinic:
        scope_parts.append(f"Clinic: {args.clinic}")
    scope_description = ", ".join(scope_parts) if scope_parts else "All Programs"

    formatter = AccessExcelFormatter(am)
    output_path = formatter.export_review_worksheet(
        reviews,
        output_path,
        scope_description=scope_description
    )

    print(f"\nExported review worksheet: {output_path}")
    print(f"  Reviews included: {len(reviews)}")
    print("\nInstructions:")
    print("  1. Open the Excel file")
    print("  2. Select a Decision for each row (Certified, Revoked, or Modified)")
    print("  3. Add reviewer notes as needed")
    print("  4. Import completed worksheet with --import-review-worksheet")


def handle_import_review_worksheet(am: AccessManager, args) -> None:
    """Import completed review worksheet from Excel."""
    from openpyxl import load_workbook

    try:
        wb = load_workbook(args.import_review_worksheet)
        ws = wb.active

        # Find the header row (look for 'Access ID')
        header_row = None
        for row in range(1, 10):
            if ws.cell(row=row, column=1).value == 'Access ID':
                header_row = row
                break

        if not header_row:
            print("Error: Could not find header row in worksheet")
            sys.exit(1)

        # Process rows
        reviewed = 0
        errors = []

        for row in range(header_row + 1, ws.max_row + 1):
            access_id = ws.cell(row=row, column=1).value
            decision = ws.cell(row=row, column=8).value  # Column H = Decision
            notes = ws.cell(row=row, column=9).value      # Column I = Notes

            if not access_id or not decision:
                continue

            try:
                am.conduct_review(
                    access_id=int(access_id),
                    reviewed_by=args.by or 'Worksheet Import',
                    status=decision,
                    notes=notes
                )
                reviewed += 1
            except Exception as e:
                errors.append(f"Row {row} (Access {access_id}): {e}")

        print(f"\nImport complete!")
        print(f"  Reviews processed: {reviewed}")

        if errors:
            print(f"  Errors: {len(errors)}")
            for error in errors[:5]:  # Show first 5 errors
                print(f"    - {error}")
            if len(errors) > 5:
                print(f"    ... and {len(errors) - 5} more")

    except Exception as e:
        print(f"Error reading worksheet: {e}")
        sys.exit(1)


def handle_assign_training(am: AccessManager, args) -> None:
    """Assign training to a user."""
    if not args.user:
        print("Error: --user required with --assign-training")
        sys.exit(1)

    if not args.training_type:
        print("Error: --training-type required with --assign-training")
        sys.exit(1)

    if not args.by:
        print("Error: --by required with --assign-training")
        sys.exit(1)

    try:
        training_id = am.assign_training(
            user_id=args.user,
            training_type=args.training_type,
            assigned_by=args.by
        )

        print(f"\nTraining assigned!")
        print(f"  Training ID: {training_id}")
        print(f"  User: {args.user}")
        print(f"  Type: {args.training_type}")
        print(f"  Assigned by: {args.by}")
    except ValueError as e:
        print(f"Error: {e}")
        sys.exit(1)


def handle_complete_training(am: AccessManager, args) -> None:
    """Mark training as completed."""
    if not args.training_id:
        print("Error: --training-id required with --complete-training")
        sys.exit(1)

    try:
        result = am.complete_training(
            training_id=args.training_id,
            completed_date=args.date,
            certificate_reference=args.certificate
        )

        print(f"\nTraining completed!")
        print(f"  Training ID: {args.training_id}")
        print(f"  User: {result['user_name']}")
        print(f"  Type: {result['training_type']}")
        print(f"  Completed: {result['completed_date']}")
        print(f"  Expires: {result['expires_date']}")
        if result.get('certificate_reference'):
            print(f"  Certificate: {result['certificate_reference']}")
    except ValueError as e:
        print(f"Error: {e}")
        sys.exit(1)


def handle_training_status(am: AccessManager, args) -> None:
    """Show training status for a user."""
    if not args.user:
        print("Error: --user required with --training-status")
        sys.exit(1)

    try:
        training = am.get_training_status(args.user)

        print(f"\nTraining status for: {args.user}")
        print("=" * 60)

        if not training:
            print("No training records found")
            return

        for record in training:
            status_marker = ""
            if record['status'] == 'Expired':
                status_marker = " [EXPIRED]"
            elif record['status'] == 'Pending':
                status_marker = " [PENDING]"

            print(f"\n{record['training_type']}{status_marker}")
            print(f"  Status: {record['status']}")
            if record.get('completed_date'):
                print(f"  Completed: {record['completed_date']}")
            if record.get('expires_date'):
                print(f"  Expires: {record['expires_date']}")
            if record.get('certificate_reference'):
                print(f"  Certificate: {record['certificate_reference']}")
    except ValueError as e:
        print(f"Error: {e}")
        sys.exit(1)


def handle_expired_training(am: AccessManager, args) -> None:
    """Show users with expired training."""
    expired = am.get_expired_training()

    print(f"\nExpired/Expiring Training")
    print("=" * 60)

    if not expired:
        print("No expired or expiring training - all training is current!")
        return

    for record in expired:
        days = int(record.get('days_until_expiry', 0))
        if days < 0:
            status = f"EXPIRED {abs(days)} days ago"
        else:
            status = f"Expires in {days} days"

        print(f"\n{record['user_name']} - {record['training_type']}")
        print(f"  Email: {record['email']}")
        print(f"  Status: {status}")
        print(f"  Expires: {record['expires_date']}")


def handle_compliance_report(am: AccessManager, args) -> None:
    """Generate a compliance report."""
    reports = ComplianceReports(am)

    report_type = args.compliance_report

    # Build filters based on arguments
    filters = {}
    if args.program:
        filters['program_id'] = args.program
    if args.clinic:
        filters['clinic_id'] = args.clinic
    if args.start_date:
        filters['start_date'] = args.start_date
    if args.end_date:
        filters['end_date'] = args.end_date

    # Generate report
    try:
        if report_type == 'access_list':
            report = reports.access_list_report(**filters)
        elif report_type == 'access_changes':
            if not args.start_date:
                print("Error: --start-date required for access_changes report")
                sys.exit(1)
            report = reports.access_changes_report(**filters)
        elif report_type == 'review_status':
            report = reports.review_status_report(**filters)
        elif report_type == 'overdue_reviews':
            report = reports.overdue_reviews_report()
        elif report_type == 'training_compliance':
            report = reports.training_compliance_report(**filters)
        elif report_type == 'terminated_audit':
            report = reports.terminated_user_audit()
        elif report_type == 'business_associates':
            report = reports.business_associate_report()
        elif report_type == 'segregation_of_duties':
            report = reports.segregation_of_duties_report(**filters)
        else:
            print(f"Error: Unknown report type: {report_type}")
            sys.exit(1)

        # Output format
        if args.output:
            # Export to Excel
            output_path = reports.export_to_excel(report_type, args.output, **filters)
            print(f"Report exported to: {output_path}")
        else:
            # Print summary to console
            print(f"\n{report_type.replace('_', ' ').title()} Report")
            print("=" * 60)
            print(f"Generated: {report.get('report_date', 'N/A')}")

            summary = report.get('summary', {})
            if summary:
                print("\nSummary:")
                for key, value in summary.items():
                    print(f"  {key.replace('_', ' ').title()}: {value}")

            # For critical reports, show status
            if 'is_compliant' in report:
                status = "PASS" if report['is_compliant'] else "FAIL"
                print(f"\nCompliance Status: {status}")

    except Exception as e:
        print(f"Error generating report: {e}")
        sys.exit(1)


# ============================================================================
# ACCESS IMPORT HANDLERS
# ============================================================================
# Handlers for bulk importing users, access, and training from Excel

def handle_import_users(am: AccessManager, args) -> None:
    """Import users from Excel file."""
    importer = AccessImporter(am)

    print(f"Importing users from: {args.import_users}")
    if args.dry_run:
        print("(Dry run - no changes will be made)")

    results = importer.import_users(args.import_users, dry_run=args.dry_run)

    print(f"\nImport Results:")
    print(f"  Imported: {results['imported']}")
    print(f"  Skipped (already exist): {results['skipped']}")
    print(f"  Errors: {len(results['errors'])}")

    if results['errors']:
        print("\nErrors:")
        for error in results['errors'][:10]:
            print(f"  - {error}")
        if len(results['errors']) > 10:
            print(f"  ... and {len(results['errors']) - 10} more")

    if results.get('user_ids') and not args.dry_run:
        print(f"\nCreated {len(results['user_ids'])} user(s)")


def handle_import_access(am: AccessManager, args) -> None:
    """Import access grants from Excel file."""
    importer = AccessImporter(am)

    print(f"Importing access grants from: {args.import_access}")
    if args.dry_run:
        print("(Dry run - no changes will be made)")

    results = importer.import_access(args.import_access, dry_run=args.dry_run)

    print(f"\nImport Results:")
    print(f"  Imported: {results['imported']}")
    print(f"  Skipped: {results['skipped']}")
    print(f"  Conflicts: {len(results.get('conflicts', []))}")
    print(f"  Errors: {len(results['errors'])}")

    if results.get('conflicts'):
        print("\nSegregation of Duties Warnings:")
        for conflict in results['conflicts'][:5]:
            print(f"  - {conflict}")
        if len(results['conflicts']) > 5:
            print(f"  ... and {len(results['conflicts']) - 5} more")

    if results['errors']:
        print("\nErrors:")
        for error in results['errors'][:10]:
            print(f"  - {error}")
        if len(results['errors']) > 10:
            print(f"  ... and {len(results['errors']) - 10} more")


def handle_import_training(am: AccessManager, args) -> None:
    """Import training records from Excel file."""
    importer = AccessImporter(am)

    print(f"Importing training records from: {args.import_training}")
    if args.dry_run:
        print("(Dry run - no changes will be made)")

    results = importer.import_training(args.import_training, dry_run=args.dry_run)

    print(f"\nImport Results:")
    print(f"  Imported: {results['imported']}")
    print(f"  Skipped: {results['skipped']}")
    print(f"  Errors: {len(results['errors'])}")

    if results['errors']:
        print("\nErrors:")
        for error in results['errors'][:10]:
            print(f"  - {error}")
        if len(results['errors']) > 10:
            print(f"  ... and {len(results['errors']) - 10} more")


def handle_import_access_template(am: AccessManager, args) -> None:
    """Import from multi-tab template."""
    importer = AccessImporter(am)

    print(f"Importing from template: {args.import_access_template}")
    if args.dry_run:
        print("(Dry run - no changes will be made)")

    results = importer.import_from_template(args.import_access_template, dry_run=args.dry_run)

    print("\n" + "=" * 60)
    print("IMPORT SUMMARY")
    print("=" * 60)

    # Users
    users = results.get('users', {})
    print(f"\nUsers:")
    print(f"  Imported: {users.get('imported', 0)}")
    print(f"  Skipped: {users.get('skipped', 0)}")
    if users.get('errors'):
        print(f"  Errors: {len(users['errors'])}")
        for error in users['errors'][:3]:
            print(f"    - {error}")

    # Access
    access = results.get('access', {})
    print(f"\nAccess Grants:")
    print(f"  Imported: {access.get('imported', 0)}")
    print(f"  Skipped: {access.get('skipped', 0)}")
    if access.get('conflicts'):
        print(f"  Conflicts: {len(access['conflicts'])}")
    if access.get('errors'):
        print(f"  Errors: {len(access['errors'])}")
        for error in access['errors'][:3]:
            print(f"    - {error}")

    # Training
    training = results.get('training', {})
    print(f"\nTraining Records:")
    print(f"  Imported: {training.get('imported', 0)}")
    print(f"  Skipped: {training.get('skipped', 0)}")
    if training.get('errors'):
        print(f"  Errors: {len(training['errors'])}")
        for error in training['errors'][:3]:
            print(f"    - {error}")

    # Total
    total_imported = (
        users.get('imported', 0) +
        access.get('imported', 0) +
        training.get('imported', 0)
    )
    total_errors = (
        len(users.get('errors', [])) +
        len(access.get('errors', [])) +
        len(training.get('errors', []))
    )

    print(f"\nTotal: {total_imported} records imported, {total_errors} errors")


def handle_generate_access_template(am: AccessManager, args) -> None:
    """Generate blank import template."""
    importer = AccessImporter(am)

    # Determine output path
    if args.output:
        output_path = args.output
    else:
        output_path = "outputs/Access_Import_Template.xlsx"

    path = importer.generate_import_template(output_path)

    print(f"\nGenerated import template: {path}")
    print("\nTemplate includes:")
    print("  - Users tab: For user information")
    print("  - Access tab: For access grants")
    print("  - Training tab: For training records")
    print("  - Instructions tab: Field definitions and valid values")
    print("\nNext steps:")
    print("  1. Open the template and fill in your data")
    print("  2. Delete the green example rows")
    print("  3. Import: python3 run.py --import-access-template '<path>'")


# ============================================================================
# MAIN ENTRY POINT
# ============================================================================

def main():
    """Main entry point for CLI."""
    parser = create_parser()
    args = parser.parse_args()

    # Initialize ConfigurationManager
    cm = ConfigurationManager(args.db)

    # Initialize AccessManager (uses same database)
    am = AccessManager(args.db)

    try:
        # ====================================================================
        # Configuration Management Commands
        # ====================================================================

        if args.init:
            handle_init(cm, args)

        elif args.import_file:
            handle_import(cm, args)

        elif args.view:
            handle_view(cm, args)

        elif args.list_programs:
            handle_list_programs(cm, args)

        elif args.audit:
            handle_audit(cm, args)

        elif args.set:
            handle_set(cm, args)

        elif args.update_provider:
            handle_update_provider(cm, args)

        elif args.export:
            handle_export(cm, args)

        elif args.compare:
            handle_compare(cm, args)

        elif args.tree:
            handle_tree(cm, args)

        elif args.create_program:
            handle_create_program(cm, args)

        elif args.create_clinic:
            handle_create_clinic(cm, args)

        elif args.create_location:
            handle_create_location(cm, args)

        # ====================================================================
        # Access Management Commands
        # ====================================================================

        elif args.init_access:
            handle_init_access(am, args)

        elif args.add_user:
            handle_add_user(am, args)

        elif args.list_users:
            handle_list_users(am, args)

        elif args.terminate_user:
            handle_terminate_user(am, args)

        elif args.grant_access:
            handle_grant_access(am, cm, args)

        elif args.revoke_access:
            handle_revoke_access(am, args)

        elif args.list_access:
            handle_list_access(am, args)

        elif args.reviews_due:
            handle_reviews_due(am, args)

        elif args.conduct_review:
            handle_conduct_review(am, args)

        elif args.export_review_worksheet:
            handle_export_review_worksheet(am, args)

        elif args.import_review_worksheet:
            handle_import_review_worksheet(am, args)

        elif args.assign_training:
            handle_assign_training(am, args)

        elif args.complete_training:
            handle_complete_training(am, args)

        elif args.training_status:
            handle_training_status(am, args)

        elif args.expired_training:
            handle_expired_training(am, args)

        elif args.compliance_report:
            handle_compliance_report(am, args)

        # ====================================================================
        # Access Import Commands
        # ====================================================================

        elif args.import_users:
            handle_import_users(am, args)

        elif args.import_access:
            handle_import_access(am, args)

        elif args.import_training:
            handle_import_training(am, args)

        elif args.import_access_template:
            handle_import_access_template(am, args)

        elif args.generate_access_template:
            handle_generate_access_template(am, args)

        else:
            parser.print_help()

    finally:
        cm.close()
        am.close()


if __name__ == "__main__":
    main()
